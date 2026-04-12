"""
Generates cbom_rules_reference.xlsx with two sheets.
Both sheets use the same 3-column layout:
    Rule ID | What | Reference
matching the table format shown in the screenshot.

Sheet 1 - "CERT-In Rules"        : rules that appear in / align with CERT-In CBOM Guidelines v2.0
Sheet 2 - "Other Standards Rules": rules sourced from NIST, RFCs, CVEs, CWEs, OWASP only
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Data — Sheet 1: CERT-In Rules
# "Reference" = specific CERT-In section / table
# ---------------------------------------------------------------------------

CERT_IN_RULES = [
    # Hash
    ("CBOM-HASH-001",    "MD5",                                  "Table 9 — Deprecated Hash Algorithm"),
    ("CBOM-HASH-002",    "SHA-1",                                "Table 9 — Deprecated Hash Algorithm"),
    # Symmetric ciphers
    ("CBOM-CIPHER-001",  "DES",                                  "Table 9 — Deprecated Cipher"),
    ("CBOM-CIPHER-002",  "3DES / TDEA",                          "Table 9 — Deprecated Cipher"),
    ("CBOM-CIPHER-003",  "RC4",                                  "Table 9 — Deprecated Cipher"),
    ("CBOM-CIPHER-004",  "RC2",                                  "Table 9 — Deprecated Cipher"),
    ("CBOM-CIPHER-005",  "IDEA",                                 "Table 9 — Deprecated Cipher"),
    ("CBOM-CIPHER-006",  "Blowfish",                             "Table 9 — Deprecated Cipher"),
    ("CBOM-CIPHER-007",  "Null / No Encryption",                 "Table 9 — Deprecated Cipher"),
    # Mode
    ("CBOM-MODE-001",    "ECB mode",                             "Insecure mode of operation"),
    # Padding
    ("CBOM-PADDING-001", "PKCS#1 v1.5 RSA padding",             "Insecure padding scheme"),
    # TLS protocols
    ("CBOM-TLS-001",     "SSLv2",                                "Deprecated TLS/SSL version"),
    ("CBOM-TLS-002",     "SSLv3",                                "Deprecated TLS/SSL version"),
    ("CBOM-TLS-003",     "TLS 1.0",                              "Deprecated TLS/SSL version"),
    ("CBOM-TLS-004",     "TLS 1.1",                              "Deprecated TLS/SSL version"),
    # Cipher suites
    ("CBOM-SUITE-001",   "Cipher suite — RC4 bulk encryption",   "Prohibited cipher suite"),
    ("CBOM-SUITE-002",   "Cipher suite — NULL encryption",       "Prohibited cipher suite"),
    ("CBOM-SUITE-003",   "Cipher suite — EXPORT grade",          "Prohibited cipher suite"),
    ("CBOM-SUITE-004",   "Cipher suite — Anonymous (no auth)",   "Prohibited cipher suite"),
    ("CBOM-SUITE-005",   "Cipher suite — DES bulk encryption",   "Prohibited cipher suite"),
    ("CBOM-SUITE-006",   "Cipher suite — MD5 MAC",               "Prohibited cipher suite"),
    ("CBOM-SUITE-007",   "Cipher suite — SHA-1 MAC",             "Deprecated cipher suite"),
    # Key sizes
    ("CBOM-KEY-001",     "RSA key < 2048 bits",                  "Minimum key size table"),
    ("CBOM-KEY-002",     "DSA key < 2048 bits",                  "Minimum key size table"),
    ("CBOM-KEY-003",     "ECC key < 256 bits",                   "Minimum key size table"),
    # Certificate signature
    ("CBOM-CERT-001",    "MD5-based certificate signature",      "Table 9 — Certificate Signature Algorithm (required CBOM element)"),
    ("CBOM-CERT-002",    "SHA-1-based certificate signature",    "Table 9 — Certificate Signature Algorithm (required CBOM element)"),
    # National algorithms
    ("CBOM-NATIONAL-002","SM2 / SM3 / SM4 (Chinese national)",  "Alignment with internationally reviewed standards (RFC 7696 §3.4)"),
    ("CBOM-NATIONAL-003","SEED / ARIA (Korean national)",        "Alignment with internationally reviewed standards (RFC 7696 §3.4)"),
    # JWT
    ("CBOM-JWT-001",     "JWT algorithm 'none' — unsigned token","Secure application development practices; token integrity"),
]

# ---------------------------------------------------------------------------
# Data — Sheet 2: Other Standards Rules
# "Reference" = primary non-CERT-In standard
# ---------------------------------------------------------------------------

OTHER_RULES = [
    # Hash (legacy / obscure)
    ("CBOM-HASH-003",    "MD4 — cryptographically broken hash",           "RFC 6150"),
    ("CBOM-HASH-004",    "MD2 — obsolete hash algorithm",                 "RFC 6149"),
    ("CBOM-HASH-005",    "SHA-224 — marginally short digest",             "NIST SP 800-131A Rev 2"),
    ("CBOM-HASH-006",    "RIPEMD-160 — non-standard aging hash",          "NIST SP 800-131A Rev 2"),
    # Key size (advisory)
    ("CBOM-KEY-004",     "AES-128 — upgrade to AES-256 advisory",         "NIST SP 800-131A Rev 2 / NIST IR 8547"),
    # Post-quantum advisories
    ("CBOM-PQC-001",     "RSA — quantum-vulnerable (Shor's algorithm)",   "NIST IR 8547"),
    ("CBOM-PQC-002",     "DSA — quantum-vulnerable (Shor's algorithm)",   "NIST IR 8547"),
    ("CBOM-PQC-003",     "ECDH / X25519 / X448 — quantum-vulnerable key exchange", "NIST IR 8547"),
    ("CBOM-PQC-004",     "ECDSA / Ed25519 / Ed448 — quantum-vulnerable signature", "NIST IR 8547"),
    ("CBOM-PQC-005",     "DH (Diffie-Hellman) — quantum-vulnerable key exchange",  "NIST IR 8547"),
    # KDF weaknesses
    ("CBOM-KDF-001",     "Weak PBE scheme (PBEWithMD5AndDES, PBEWithSHA1AndRC2)", "CWE-1240 / NIST SP 800-131A Rev 2"),
    ("CBOM-KDF-002",     "PBKDF2 with SHA-1 or MD5 as PRF",              "NIST SP 800-132 / RFC 8018"),
    # RNG
    ("CBOM-RNG-001",     "SHA1PRNG — weak Java PRNG",                    "CWE-338 / NIST SP 800-131A Rev 2"),
    ("CBOM-RNG-003",     "passlib weak password hash (MD5-crypt, SHA1-crypt, DES-crypt)", "NIST SP 800-63B §5.1.1 / CWE-916"),
    # Misconfiguration
    ("CBOM-NULL-001",    "NullCipher — no encryption performed",          "CWE-1240 / CWE-327"),
    ("CBOM-IV-001",      "Hardcoded zero IV (new byte[16])",              "CWE-1240 / CWE-329"),
    ("CBOM-KS-001",      "Deprecated JKS keystore format",               "JEP 229 / NIST SP 800-131A Rev 2"),
    # National algorithms (non-CERT-In)
    ("CBOM-NATIONAL-001","GOST — Russian national cryptographic algorithm","RFC 7696 §3.4"),
    # JWT
    ("CBOM-JWT-002",     "JWT HMAC algorithm — symmetric signing risk",   "RFC 7519 / OWASP JWT Security Cheat Sheet"),
]

# ---------------------------------------------------------------------------
# Styles
# ---------------------------------------------------------------------------

CERT_IN_HEADER_FILL = PatternFill("solid", fgColor="1F4E79")   # dark navy blue
CERT_IN_ROW_A       = PatternFill("solid", fgColor="FFFFFF")   # white
CERT_IN_ROW_B       = PatternFill("solid", fgColor="DDEEFF")   # light blue
OTHER_HEADER_FILL   = PatternFill("solid", fgColor="375623")   # dark green
OTHER_ROW_A         = PatternFill("solid", fgColor="FFFFFF")   # white
OTHER_ROW_B         = PatternFill("solid", fgColor="E2EFDA")   # light green

HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=12)
BODY_FONT   = Font(name="Calibri", size=11)
ID_FONT     = Font(name="Calibri", bold=True, size=11)

WRAP_MID    = Alignment(wrap_text=True, vertical="center")
WRAP_TOP    = Alignment(wrap_text=True, vertical="center")

thin = Side(style="thin", color="BBBBBB")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)


def make_sheet(wb, title, headers, rows, header_fill, row_a, row_b, col_widths):
    ws = wb.create_sheet(title)

    # Header row
    ws.append(headers)
    ws.row_dimensions[1].height = 28
    for ci in range(1, len(headers) + 1):
        c = ws.cell(row=1, column=ci)
        c.font      = HEADER_FONT
        c.fill      = header_fill
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border    = BORDER

    # Data rows
    for ri, row in enumerate(rows, start=2):
        fill = row_b if ri % 2 == 0 else row_a
        ws.row_dimensions[ri].height = 22
        for ci, val in enumerate(row, start=1):
            c = ws.cell(row=ri, column=ci)
            c.value     = val
            c.fill      = fill
            c.alignment = WRAP_MID
            c.border    = BORDER
            if ci == 1:                      # Rule ID column — bold
                c.font = ID_FONT
            else:
                c.font = BODY_FONT

    # Column widths
    for ci, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"
    return ws


# ---------------------------------------------------------------------------
# Build workbook
# ---------------------------------------------------------------------------

wb = openpyxl.Workbook()
wb.remove(wb.active)   # remove default empty sheet

make_sheet(
    wb,
    title        = "CERT-In Rules",
    headers      = ["Rule ID", "What", "CERT-In Reference"],
    rows         = CERT_IN_RULES,
    header_fill  = CERT_IN_HEADER_FILL,
    row_a        = CERT_IN_ROW_A,
    row_b        = CERT_IN_ROW_B,
    col_widths   = [18, 46, 62],
)

make_sheet(
    wb,
    title        = "Other Standards Rules",
    headers      = ["Rule ID", "What", "Reference"],
    rows         = OTHER_RULES,
    header_fill  = OTHER_HEADER_FILL,
    row_a        = OTHER_ROW_A,
    row_b        = OTHER_ROW_B,
    col_widths   = [18, 56, 52],
)

# ---------------------------------------------------------------------------
# Save
# ---------------------------------------------------------------------------

out = r"c:\Users\Nochu\OneDrive\cbomscanner\cbom-scanner\cbom_rules_reference.xlsx"
wb.save(out)
print(f"Saved: {out}")
print(f"  Sheet 1 'CERT-In Rules'         : {len(CERT_IN_RULES)} rules")
print(f"  Sheet 2 'Other Standards Rules' : {len(OTHER_RULES)} rules")
