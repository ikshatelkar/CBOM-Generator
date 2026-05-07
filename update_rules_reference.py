"""
Generates the cbom_rules_reference.xlsx with ALL current detection rules
extracted from every Go rule source file.
"""
import re
import os
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUTPUT_FILE = r"c:\Users\Nochu\OneDrive\cbomscanner\cbom-scanner\cbom_rules_reference.xlsx"

BASE = r"c:\Users\Nochu\OneDrive\cbomscanner\cbom-scanner\pkg\rules"

RULE_FILES = [
    (os.path.join(BASE, "golang",     "stdlib.go"),         "Go"),
    (os.path.join(BASE, "java",       "jca.go"),            "Java"),
    (os.path.join(BASE, "java",       "commons.go"),        "Java"),
    (os.path.join(BASE, "java",       "spring.go"),         "Java"),
    (os.path.join(BASE, "java",       "bouncycastle.go"),   "Java"),
    (os.path.join(BASE, "python",     "pyca.go"),           "Python"),
    (os.path.join(BASE, "php",        "rules.go"),          "PHP"),
    (os.path.join(BASE, "ruby",       "rules.go"),          "Ruby"),
    (os.path.join(BASE, "flutter",    "cryptography.go"),   "Dart/Flutter"),
    (os.path.join(BASE, "csharp",     "rules.go"),          "C#"),
    (os.path.join(BASE, "javascript", "rules.go"),          "JavaScript/TypeScript"),
    (os.path.join(BASE, "rust",       "rules.go"),          "Rust"),
]

# ── Helpers to infer Asset Name and Asset Type from a rule ID ─────────────────

def infer_asset(rule_id, bundle):
    """Return (asset_name, asset_type) from a rule ID + bundle string."""
    rid = rule_id.lower()

    # --- Asset Type mapping based on keywords in id --------------------------
    type_map = [
        (["gcm", "ccm", "siv", "ocb", "poly1305", "chacha20poly1305",
          "xchacha20", "aesgcm", "aesocb", "chachapoly", "aead",
          "nacl-box", "nacl-secretbox"],               "AEAD Cipher"),
        (["cbc", "cfb", "ofb", "ctr", "xts", "cipher-mode"],
                                                        "Block Cipher Mode"),
        (["aes", "des", "rc2", "rc4", "rc5", "blowfish",
          "twofish", "camellia", "idea", "salsa",
          "chacha20-stream", "stream"],                 "Symmetric Cipher"),
        (["sha3", "sha256", "sha512", "sha1", "md5",
          "blake2", "ripemd", "whirlpool", "hash",
          "digest", "shake"],                           "Hash Function"),
        (["hmac", "cmac", "poly1305", "mac", "tag",
          "totp", "hotp"],                              "MAC / OTP"),
        (["pbkdf2", "scrypt", "argon2", "bcrypt",
          "passwor", "pwhash", "kdf"],                  "Key Derivation / Password Hash"),
        (["hkdf", "derive", "kdf"],                     "Key Derivation"),
        (["rsa"],                                        "RSA Asymmetric"),
        (["ecdsa", "ecdh", "elliptic", "curve", "ec-key",
          "p256", "p384", "p521", "secp"],               "Elliptic Curve"),
        (["ed25519", "ed448", "eddsa"],                  "EdDSA Signature"),
        (["x25519", "x448", "dh-compute", "key-agree",
          "scalarmult"],                                 "Key Agreement (ECDH/DH)"),
        (["dsa"],                                        "DSA Signature"),
        (["mlkem", "ml-kem", "kyber", "kem"],            "Post-Quantum KEM"),
        (["mldsa", "ml-dsa", "slhdsa", "slh-dsa",
          "sphincs", "dilithium"],                       "Post-Quantum Signature"),
        (["jwt", "jws", "token"],                        "JWT / Token Signing"),
        (["tls", "ssl", "https", "x509", "cert",
          "pkcs", "pem", "keypair"],                     "TLS / Certificate"),
        (["rand", "prng", "random", "uuid"],             "Random / PRNG"),
        (["shamir"],                                     "Secret Sharing"),
        (["sign", "verify", "signature"],                "Digital Signature"),
        (["encrypt", "decrypt"],                         "Encryption / Decryption"),
        (["key-gen", "keygen", "generatekey",
          "generate-key", "newkey"],                     "Key Generation"),
        (["openssl"],                                     "OpenSSL Wrapper"),
        (["sodium", "libsodium"],                        "Libsodium"),
        (["bcrypt"],                                     "Password Hash (bcrypt)"),
    ]

    asset_type = "Cryptographic Primitive"
    for keywords, atype in type_map:
        if any(k in rid for k in keywords):
            asset_type = atype
            break

    # --- Asset Name: tidy up the rule ID ------------------------------------
    name = rule_id
    # strip language prefix (go-, java-, py-, php-, rb-, flutter-, cs-, js-, rs-)
    for pfx in ["go-", "java-", "py-", "pyca-", "php-", "rb-",
                "flutter-", "cs-", "js-", "rs-", "ruby-"]:
        if name.startswith(pfx):
            name = name[len(pfx):]
            break
    name = name.replace("-", " ").replace("_", " ").title()
    return name, asset_type


def extract_rules(filepath, language):
    """Parse a Go source file and return list of (rule_id, bundle, language, asset_name, asset_type)."""
    try:
        src = open(filepath, encoding="utf-8").read()
    except FileNotFoundError:
        return []

    # Extract Rule struct literals: find all ID: "..." and Bundle: "..." pairs
    # They appear together inside each &detection.Rule{ ... } block
    results = []

    # Split on rule struct starts — each Rule block ends at the next closing brace
    # Strategy: find every occurrence of `ID: "..."` and the nearest `Bundle: "..."`
    id_pattern    = re.compile(r'ID:\s+"([^"]+)"')
    bundle_pattern = re.compile(r'Bundle:\s+"([^"]+)"')

    # Find all IDs with their position
    id_matches = [(m.group(1), m.start()) for m in id_pattern.finditer(src)]
    bundle_matches = [(m.group(1), m.start()) for m in bundle_pattern.finditer(src)]

    for rule_id, id_pos in id_matches:
        # Find the nearest bundle that comes after this id (within ~500 chars)
        bundle = "Unknown"
        for b, b_pos in bundle_matches:
            if b_pos > id_pos and b_pos - id_pos < 500:
                bundle = b
                break

        asset_name, asset_type = infer_asset(rule_id, bundle)
        results.append((rule_id, bundle, language, asset_name, asset_type))

    return results


# ── Collect all rules ─────────────────────────────────────────────────────────

all_rules = []
for filepath, lang in RULE_FILES:
    rules = extract_rules(filepath, lang)
    all_rules.extend(rules)
    print(f"  {lang:25s} ({os.path.basename(filepath):20s}): {len(rules)} rules")

print(f"\nTotal rules: {len(all_rules)}\n")

# ── Colour palette ────────────────────────────────────────────────────────────
CLR_HEADER_BG  = "1F3864"
CLR_HEADER_FG  = "FFFFFF"
CLR_ROW_ALT    = "EEF2FA"
CLR_ROW_EVEN   = "FFFFFF"
CLR_BORDER     = "B8CCE4"
CLR_ACCENT     = "2E75B6"

LANG_COLORS = {
    "Go":                    ("DCEEFB", "1A4971"),
    "Java":                  ("FFF0DC", "7F4E00"),
    "Python":                ("E8F5E9", "1B5E20"),
    "PHP":                   ("F3E5F5", "4A148C"),
    "Ruby":                  ("FCE4EC", "880E4F"),
    "Dart/Flutter":          ("E0F7FA", "006064"),
    "C#":                    ("EDE7F6", "311B92"),
    "JavaScript/TypeScript": ("FFFDE7", "F57F17"),
    "Rust":                  ("FBE9E7", "BF360C"),
}

def thin_border():
    s = Side(style="thin", color=CLR_BORDER)
    return Border(left=s, right=s, top=s, bottom=s)

# ── Build workbook ────────────────────────────────────────────────────────────
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "All Rules"

HEADERS = ["#", "Rule ID", "Language", "Bundle / Library", "Asset Name", "Asset Type"]
COL_WIDTHS = [5, 38, 22, 22, 30, 28]

# Title
ws.merge_cells("A1:F1")
tc = ws["A1"]
tc.value = "CBOM Scanner — Detection Rules Reference"
tc.font = Font(name="Calibri", size=13, bold=True, color=CLR_HEADER_FG)
tc.fill = PatternFill("solid", fgColor=CLR_HEADER_BG)
tc.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 26

# Subtitle
ws.merge_cells("A2:F2")
sc = ws["A2"]
sc.value = f"Auto-generated from Go source rules  |  Total rules: {len(all_rules)}  |  Languages: {len(RULE_FILES)}"
sc.font = Font(name="Calibri", size=9, italic=True, color="595959")
sc.fill = PatternFill("solid", fgColor="D9E2F3")
sc.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[2].height = 16

ws.row_dimensions[3].height = 5  # spacer

# Header row
for ci, (h, w) in enumerate(zip(HEADERS, COL_WIDTHS), 1):
    cell = ws.cell(row=4, column=ci, value=h)
    cell.font = Font(name="Calibri", size=10, bold=True, color=CLR_HEADER_FG)
    cell.fill = PatternFill("solid", fgColor=CLR_HEADER_BG)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = thin_border()
    ws.column_dimensions[get_column_letter(ci)].width = w
ws.row_dimensions[4].height = 20

# Data rows
for i, (rule_id, bundle, language, asset_name, asset_type) in enumerate(all_rules, 1):
    row = 4 + i
    bg, fg = LANG_COLORS.get(language, (CLR_ROW_ALT if i % 2 == 0 else CLR_ROW_EVEN, "000000"))
    vals = [i, rule_id, language, bundle, asset_name, asset_type]
    for ci, v in enumerate(vals, 1):
        cell = ws.cell(row=row, column=ci, value=v)
        cell.fill = PatternFill("solid", fgColor=bg)
        cell.border = thin_border()
        cell.font = Font(name="Calibri", size=9,
                         bold=(ci == 2),
                         color=(CLR_ACCENT if ci == 2 else fg))
        cell.alignment = Alignment(
            vertical="center",
            horizontal="center" if ci in (1, 5) else "left"
        )
    ws.row_dimensions[row].height = 16

ws.freeze_panes = "A5"
ws.auto_filter.ref = "A4:F4"

# ── Per-language summary sheet ────────────────────────────────────────────────
ws2 = wb.create_sheet("Summary by Language")
ws2.column_dimensions["A"].width = 26
ws2.column_dimensions["B"].width = 14
ws2.column_dimensions["C"].width = 40

# header
for ci, h in enumerate(["Language", "Rule Count", "Bundles"], 1):
    c = ws2.cell(row=1, column=ci, value=h)
    c.font = Font(name="Calibri", size=10, bold=True, color=CLR_HEADER_FG)
    c.fill = PatternFill("solid", fgColor=CLR_HEADER_BG)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = thin_border()
ws2.row_dimensions[1].height = 20

from collections import defaultdict
lang_stats = defaultdict(list)
for rule_id, bundle, language, _, _ in all_rules:
    lang_stats[language].append(bundle)

for ri, (lang, bundles) in enumerate(sorted(lang_stats.items()), 2):
    unique_bundles = sorted(set(bundles))
    bg, fg = LANG_COLORS.get(lang, ("FFFFFF", "000000"))
    row_vals = [lang, len(bundles), ", ".join(unique_bundles)]
    for ci, v in enumerate(row_vals, 1):
        c = ws2.cell(row=ri, column=ci, value=v)
        c.font = Font(name="Calibri", size=9, bold=(ci == 1))
        c.fill = PatternFill("solid", fgColor=bg)
        c.border = thin_border()
        c.alignment = Alignment(vertical="center",
                                horizontal="center" if ci == 2 else "left",
                                wrap_text=(ci == 3))
    ws2.row_dimensions[ri].height = 18

# ── Per-language detail sheets ────────────────────────────────────────────────
for lang, (bg, fg) in LANG_COLORS.items():
    lang_rules = [(rid, bnd, an, at) for rid, bnd, lg, an, at in all_rules if lg == lang]
    if not lang_rules:
        continue
    safe_name = lang[:28].replace("/", "_")
    ws_lang = wb.create_sheet(safe_name)
    ws_lang.column_dimensions["A"].width = 5
    ws_lang.column_dimensions["B"].width = 40
    ws_lang.column_dimensions["C"].width = 22
    ws_lang.column_dimensions["D"].width = 30
    ws_lang.column_dimensions["E"].width = 28

    # title
    ws_lang.merge_cells("A1:E1")
    t = ws_lang["A1"]
    t.value = f"{lang} — {len(lang_rules)} Detection Rules"
    t.font = Font(name="Calibri", size=12, bold=True, color=CLR_HEADER_FG)
    t.fill = PatternFill("solid", fgColor=CLR_HEADER_BG)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws_lang.row_dimensions[1].height = 22

    for ci, h in enumerate(["#", "Rule ID", "Bundle", "Asset Name", "Asset Type"], 1):
        c = ws_lang.cell(row=2, column=ci, value=h)
        c.font = Font(name="Calibri", size=9, bold=True, color=CLR_HEADER_FG)
        c.fill = PatternFill("solid", fgColor=CLR_ACCENT)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = thin_border()
    ws_lang.row_dimensions[2].height = 18

    for ri, (rid, bnd, an, at) in enumerate(lang_rules, 1):
        row_bg = bg if ri % 2 == 0 else "FFFFFF"
        for ci, v in enumerate([ri, rid, bnd, an, at], 1):
            c = ws_lang.cell(row=2 + ri, column=ci, value=v)
            c.font = Font(name="Calibri", size=9, bold=(ci == 2), color=(CLR_ACCENT if ci == 2 else "000000"))
            c.fill = PatternFill("solid", fgColor=row_bg)
            c.border = thin_border()
            c.alignment = Alignment(vertical="center",
                                    horizontal="center" if ci == 1 else "left")
        ws_lang.row_dimensions[2 + ri].height = 15

    ws_lang.freeze_panes = "A3"

wb.save(OUTPUT_FILE)
print(f"Saved -> {OUTPUT_FILE}")
