"""
generate_rules_reference.py
Generates an editable Excel workbook with one sheet per language,
listing every detection rule in the cbom-scanner.
"""

import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
import os

# ---------------------------------------------------------------------------
# Rule data: (rule_id, library/bundle, what_it_detects, code_pattern)
# ---------------------------------------------------------------------------

RULES = {

    # =========================================================================
    # Python
    # =========================================================================
    "Python": [
        # pyca/cryptography
        ("pyca-symmetric-algorithm",    "pyca/cryptography",  "Symmetric cipher algorithm constructor",               "algorithms.AES/TripleDES/Camellia/CAST5/SEED/SM4/Blowfish/IDEA/ARC4/ChaCha20(key)"),
        ("pyca-cipher-mode",            "pyca/cryptography",  "Block cipher mode of operation",                       "modes.CBC/CTR/OFB/CFB/GCM/XTS/ECB/SIV/CCM(iv)"),
        ("pyca-aead",                   "pyca/cryptography",  "AEAD cipher (authenticated encryption)",               "AESGCM/AESCCM/AESGCMSIV/AESSIV/AESOCB3/ChaCha20Poly1305/XChaCha20Poly1305(key)"),
        ("pyca-hash",                   "pyca/cryptography",  "Hash algorithm constructor",                           "hashes.SHA1/SHA256/SHA3_256/MD5/BLAKE2b/SM3/SHAKE128/...()"),
        ("pyca-hmac",                   "pyca/cryptography",  "HMAC message authentication code",                    "hmac.HMAC(key, algorithm, ...)"),
        ("pyca-cmac",                   "pyca/cryptography",  "CMAC message authentication code",                    "CMAC(algorithm())"),
        ("pyca-rsa-keygen",             "pyca/cryptography",  "RSA private key generation",                           "rsa.generate_private_key(...)"),
        ("pyca-ec-keygen",              "pyca/cryptography",  "Elliptic curve (EC) private key generation",           "ec.generate_private_key(SECP256R1/SECP384R1/..., ...)"),
        ("pyca-dsa-keygen",             "pyca/cryptography",  "DSA private key generation",                           "dsa.generate_private_key(key_size, ...)"),
        ("pyca-dh-keygen",              "pyca/cryptography",  "Diffie-Hellman private key generation",                "parameters.generate_private_key()"),
        ("pyca-ed25519",                "pyca/cryptography",  "Ed25519 signing key generation",                       "ed25519.Ed25519PrivateKey.generate()"),
        ("pyca-ed448",                  "pyca/cryptography",  "Ed448 signing key generation",                         "ed448.Ed448PrivateKey.generate()"),
        ("pyca-x25519",                 "pyca/cryptography",  "X25519 key exchange key generation",                   "x25519.X25519PrivateKey.generate()"),
        ("pyca-x448",                   "pyca/cryptography",  "X448 key exchange key generation",                     "x448.X448PrivateKey.generate()"),
        ("pyca-kdf",                    "pyca/cryptography",  "Key derivation function (PBKDF2, Scrypt, HKDF, etc.)", "PBKDF2HMAC/Scrypt/HKDF/ConcatKDF/X963KDF/KBKDFHMAC(...)"),
        ("pyca-fernet",                 "pyca/cryptography",  "Fernet symmetric authenticated encryption",            "Fernet(key) / MultiFernet([...])"),
        ("pyca-rsa-sign",               "pyca/cryptography",  "RSA signing operation",                                "private_key.sign(data, padding.PSS/PKCS1v15, hash)"),
        ("pyca-ec-sign",                "pyca/cryptography",  "EC signing operation",                                 "private_key.sign(data, ec.ECDSA(hash))"),
        ("pyca-rsa-encrypt",            "pyca/cryptography",  "RSA encryption (public key)",                          "public_key.encrypt(data, padding.OAEP/PKCS1v15)"),
        ("pyca-ciphersuite",            "pyca/cryptography",  "TLS cipher suite string configuration",                "context.set_ciphers('ECDHE-RSA-AES256-GCM-SHA384:...')"),
        ("pyca-ssl-context",            "pyca/cryptography",  "SSL/TLS context creation",                             "ssl.SSLContext(...) / ssl.create_default_context()"),
        ("pyca-os-urandom",             "pyca/cryptography",  "OS-level CSPRNG",                                      "os.urandom(n)"),
        ("pyca-secrets",                "pyca/cryptography",  "Secrets module CSPRNG",                                "secrets.token_bytes/token_hex/SystemRandom()"),
        ("pyca-blake3",                 "pyca/cryptography",  "BLAKE3 hash",                                          "blake3.hash(...) / blake3.Blake3(...)"),
        ("liboqs-kem",                  "liboqs-python",      "Post-quantum Key Encapsulation Mechanism",             "oqs.KeyEncapsulation('Kyber768/BIKE/...')"),
        ("liboqs-signature",            "liboqs-python",      "Post-quantum digital signature",                       "oqs.Signature('Dilithium3/FALCON/SPHINCS+/...')"),
        ("pyca-mldsa",                  "pyca/cryptography",  "ML-DSA (FIPS 204 / Dilithium) key generation",         "MLDSAPrivateKey.generate(MLDSAParameters.MLDSA44/65/87)"),
        ("pyca-mlkem",                  "pyca/cryptography",  "ML-KEM (FIPS 203 / Kyber) key generation",             "MLKEMPrivateKey.generate(MLKEMParameters.MLKEM512/768/1024)"),
        ("pyca-argon2",                 "pyca/cryptography",  "Argon2 password hashing (native pyca)",                "Argon2id/Argon2i/Argon2d(...).hash(password)"),
        ("pyca-poly1305",               "pyca/cryptography",  "Poly1305 MAC",                                         "poly1305.Poly1305(key)"),
        ("pyca-ecdh-exchange",          "pyca/cryptography",  "ECDH key exchange",                                    "private_key.exchange(ec.ECDH(), peer_public_key)"),
        ("pyca-bcrypt",                 "pyca/cryptography",  "bcrypt KDF (native pyca)",                             "bcrypt(password, salt, rounds, backend)"),
        ("pyca-rsa-decrypt",            "pyca/cryptography",  "RSA decryption (private key)",                         "private_key.decrypt(ciphertext, padding.OAEP/PKCS1v15)"),
        # stdlib
        ("py-hashlib-algo",             "hashlib",            "hashlib hash algorithm",                               "hashlib.sha256() / hashlib.md5() / hashlib.sha3_256()"),
        ("py-hashlib-new",              "hashlib",            "hashlib.new() generic hash",                           "hashlib.new('sha256', data)"),
        ("py-hashlib-pbkdf2",           "hashlib",            "PBKDF2-HMAC key derivation",                           "hashlib.pbkdf2_hmac('sha256', password, salt, iterations)"),
        ("py-hmac-new",                 "hmac",               "HMAC via stdlib hmac.new()",                           "hmac.new(key, msg, digestmod=hashlib.sha256)"),
        # PyCryptodome
        ("pycryptodome-aes-new",        "PyCryptodome",       "AES cipher (PyCryptodome)",                            "AES.new(key, AES.MODE_CBC, iv)"),
        ("pycryptodome-symmetric-new",  "PyCryptodome",       "Symmetric cipher (PyCryptodome)",                      "DES/DES3/Blowfish/ARC4/ChaCha20/ARC2.new(key, ...)"),
        ("pycryptodome-hash-new",       "PyCryptodome",       "Hash algorithm (PyCryptodome)",                        "MD5/SHA256/SHA1/SHA3_256.new()"),
        ("pycryptodome-rsa-generate",   "PyCryptodome",       "RSA key generation (PyCryptodome)",                    "RSA.generate(2048)"),
        ("pycryptodome-ecc-generate",   "PyCryptodome",       "ECC key generation (PyCryptodome)",                    "ECC.generate(curve='P-256')"),
        ("pycryptodome-rsa-pkcs",       "PyCryptodome",       "RSA OAEP/PKCS#1 v1.5 encryption (PyCryptodome)",       "PKCS1_OAEP.new(key) / PKCS1_v1_5.new(key)"),
        ("pycryptodome-hmac",           "PyCryptodome",       "HMAC (PyCryptodome)",                                  "HMAC.new(key, digestmod=SHA256)"),
        ("pycryptodome-kdf",            "PyCryptodome",       "KDF: PBKDF2 / scrypt / bcrypt (PyCryptodome)",         "PBKDF2(password, salt, ...) / scrypt(...) / bcrypt(...)"),
        # PyNaCl
        ("pynacl-signing-key",          "PyNaCl",             "Ed25519 signing key (libsodium)",                      "nacl.signing.SigningKey.generate()"),
        ("pynacl-secret-box",           "PyNaCl",             "XSalsa20-Poly1305 secret-key encryption",              "nacl.secret.SecretBox(key)"),
        ("pynacl-public-box",           "PyNaCl",             "Curve25519 + XSalsa20-Poly1305 public-key box",        "nacl.public.Box(private_key, public_key)"),
        # PyJWT
        ("pyjwt-encode",                "PyJWT",              "JWT encode (token creation)",                          "jwt.encode(payload, key, algorithm='HS256')"),
        ("pyjwt-decode",                "PyJWT",              "JWT decode (token verification)",                      "jwt.decode(token, key, algorithms=['RS256'])"),
        # passlib
        ("passlib-hash-using",          "passlib",            "passlib hash scheme (argon2, bcrypt, sha512_crypt…)",   "passlib.hash.<scheme>.using(...).hash(password)"),
        ("passlib-import",              "passlib",            "passlib hash import (md5_crypt, sha256_crypt, etc.)",   "from passlib.hash import md5_crypt"),
        # paramiko
        ("paramiko-rsa-key",            "paramiko",           "RSA key generation (SSH)",                             "paramiko.RSAKey.generate(bits=2048)"),
        ("paramiko-ecdsa-key",          "paramiko",           "ECDSA key generation (SSH)",                           "paramiko.ECDSAKey.generate(bits=256)"),
        ("paramiko-dss-key",            "paramiko",           "DSS/DSA key generation (SSH)",                         "paramiko.DSSKey.generate(bits=1024)"),
        ("paramiko-transport-ciphers",  "paramiko",           "SSH transport cipher/MAC configuration",               "transport.get_security_options().ciphers = [...]"),
        # gap-fill
        ("pyca-tripledes-bare",         "pyca/cryptography",  "TripleDES bare import usage",                          "TripleDES(key)  [after 'from algorithms import TripleDES']"),
        ("py-dh-kex-class",             "paramiko",           "Diffie-Hellman KEX class (SSH)",                       "class KexGroup1/KexGroup14/KexGroup16  name='diffie-hellman-...'"),
        ("pyca-dh-hazmat",              "pyca/cryptography",  "DH hazmat API (parameters / key generation)",          "dh.generate_parameters() / dh.DHParameterNumbers(...)"),
        ("py-stdlib-hmac-bare",         "hmac",               "Bare HMAC call after from-import",                     "HMAC(key, msg, digest)  [after 'from hmac import HMAC']"),
        ("py-os-urandom",               "os",                 "os.urandom CSPRNG (bare call)",                        "os.urandom(n)"),
        ("py-hashlib-bare-call",        "hashlib",            "Bare hashlib function call after from-import",         "md5(...) / sha1(...) / sha256(...)  [after 'from hashlib import ...']"),
        ("pyca-hash-reference",         "pyca/cryptography",  "Hash class reference (no parentheses)",                "hashes.SHA1 / hashes.SHA256  [used as algorithm reference]"),
        ("pynacl-verify-key",           "PyNaCl",             "Ed25519 verify key",                                   "nacl.signing.VerifyKey(bytes)"),
        ("pyca-ecdsa-sign",             "pyca/cryptography",  "ECDSA signing algorithm reference",                    "ec.ECDSA(hashes.SHA256())"),
        ("pyca-x25519-pubkey",          "pyca/cryptography",  "X25519 public key from bytes",                         "X25519PublicKey.from_public_bytes(data)"),
        ("pyca-pkcs1v15",               "pyca/cryptography",  "PKCS#1 v1.5 padding reference",                        "padding.PKCS1v15()"),
        ("pyca-pkcs7",                  "pyca/cryptography",  "PKCS#7 block cipher padding",                          "padding.PKCS7(block_size)"),
        ("pyca-ec-curve-standalone",    "pyca/cryptography",  "EC named curve standalone object",                     "ec.SECP256R1() / ec.SECP384R1() / ec.SECP521R1()"),
        ("pyca-mode-reference",         "pyca/cryptography",  "Cipher mode class reference (no parentheses)",         "modes.CBC / modes.CTR  [used in dict/config]"),
    ],

    # =========================================================================
    # Go
    # =========================================================================
    "Go": [
        # stdlib - symmetric
        ("go-aes-newcipher",            "crypto/aes",         "AES block cipher",                                     "aes.NewCipher(key)"),
        ("go-des-newcipher",            "crypto/des",         "DES block cipher",                                     "des.NewCipher(key)"),
        ("go-tripledes-newcipher",      "crypto/des",         "TripleDES (3DES) block cipher",                        "des.NewTripleDESCipher(key)"),
        ("go-rc4-newcipher",            "crypto/rc4",         "RC4 stream cipher",                                    "rc4.NewCipher(key)"),
        # cipher modes
        ("go-cipher-gcm",               "crypto/cipher",      "AES-GCM AEAD mode",                                    "cipher.NewGCM(block)"),
        ("go-cipher-cbc",               "crypto/cipher",      "CBC block cipher mode",                                "cipher.NewCBCEncrypter/Decrypter(block, iv)"),
        ("go-cipher-cfb",               "crypto/cipher",      "CFB block cipher mode",                                "cipher.NewCFBEncrypter/Decrypter(block, iv)"),
        ("go-cipher-ofb",               "crypto/cipher",      "OFB block cipher mode",                                "cipher.NewOFB(block, iv)"),
        ("go-cipher-ctr",               "crypto/cipher",      "CTR block cipher mode",                                "cipher.NewCTR(block, iv)"),
        # hash
        ("go-md5",                      "crypto/md5",         "MD5 hash",                                             "md5.New() / md5.Sum(data)"),
        ("go-sha1",                     "crypto/sha1",        "SHA-1 hash",                                           "sha1.New() / sha1.Sum(data)"),
        ("go-sha256",                   "crypto/sha256",      "SHA-256 / SHA-224 hash",                               "sha256.New() / sha256.Sum256(data)"),
        ("go-sha512",                   "crypto/sha512",      "SHA-512 / SHA-384 / SHA-512/256 hash",                 "sha512.New() / sha512.Sum512(data)"),
        ("go-sha3",                     "golang.org/x/crypto","SHA-3 hash (Keccak)",                                  "sha3.New256/New384/New512/NewLegacyKeccak256()"),
        # HMAC
        ("go-hmac-new",                 "crypto/hmac",        "HMAC message authentication code",                    "hmac.New(sha256.New, key)"),
        # RSA
        ("go-rsa-generatekey",          "crypto/rsa",         "RSA key pair generation",                              "rsa.GenerateKey(rand, bits)"),
        ("go-rsa-encrypt",              "crypto/rsa",         "RSA encryption (OAEP / PKCS#1 v1.5)",                  "rsa.EncryptOAEP/EncryptPKCS1v15(...)"),
        ("go-rsa-decrypt",              "crypto/rsa",         "RSA decryption",                                       "rsa.DecryptOAEP/DecryptPKCS1v15(...)"),
        ("go-rsa-sign",                 "crypto/rsa",         "RSA signing (PSS / PKCS#1 v1.5)",                      "rsa.SignPSS/SignPKCS1v15(...)"),
        ("go-rsa-verify",               "crypto/rsa",         "RSA signature verification",                           "rsa.VerifyPSS/VerifyPKCS1v15(...)"),
        # ECDSA
        ("go-ecdsa-generatekey",        "crypto/ecdsa",       "ECDSA key pair generation",                            "ecdsa.GenerateKey(elliptic.P256(), rand)"),
        ("go-ecdsa-sign",               "crypto/ecdsa",       "ECDSA signing",                                        "ecdsa.Sign(rand, privateKey, hash)"),
        ("go-ecdsa-verify",             "crypto/ecdsa",       "ECDSA signature verification",                         "ecdsa.Verify(publicKey, hash, r, s)"),
        ("go-elliptic-curve",           "crypto/elliptic",    "Elliptic curve reference",                             "elliptic.P256/P384/P521/P224()"),
        # ECDH
        ("go-ecdh-generatekey",         "crypto/ecdh",        "ECDH key generation (Go 1.20+)",                       "ecdh.P256/P384/P521/X25519().GenerateKey(rand)"),
        # Ed25519
        ("go-ed25519-generatekey",      "crypto/ed25519",     "Ed25519 key pair generation",                          "ed25519.GenerateKey(rand)"),
        ("go-ed25519-sign",             "crypto/ed25519",     "Ed25519 signing",                                      "ed25519.Sign(privateKey, message)"),
        ("go-ed25519-verify",           "crypto/ed25519",     "Ed25519 signature verification",                       "ed25519.Verify(publicKey, message, sig)"),
        # DSA
        ("go-dsa-generatekey",          "crypto/dsa",         "DSA key pair generation (deprecated)",                 "dsa.GenerateKey(&key, rand)"),
        # CSPRNG
        ("go-crypto-rand-read",         "crypto/rand",        "Cryptographically secure random bytes",                "rand.Read(buf) / rand.Int(rand, max)"),
        # TLS
        ("go-tls-minversion",           "crypto/tls",         "TLS minimum version configuration",                    "tls.Config{MinVersion: tls.VersionTLS12}"),
        ("go-tls-dial",                 "crypto/tls",         "TLS client connection",                                "tls.Dial(network, addr, config)"),
        ("go-tls-listen",               "crypto/tls",         "TLS server listener",                                  "tls.Listen(network, addr, config)"),
        ("go-tls-ciphersuites",         "crypto/tls",         "TLS cipher suite selection",                           "tls.Config{CipherSuites: [...]}"),
        ("go-tls-curve-preferences",    "crypto/tls",         "TLS curve preferences",                                "tls.Config{CurvePreferences: [tls.CurveP256, ...]}"),
        ("go-tls-x25519mlkem768",       "crypto/tls",         "Post-quantum TLS hybrid KEM (Go 1.23+)",               "tls.X25519MLKEM768 curve preference"),
        # x/crypto
        ("go-xcrypto-chacha20poly1305", "golang.org/x/crypto","ChaCha20-Poly1305 AEAD",                               "chacha20poly1305.New(key)"),
        ("go-xcrypto-xchacha20poly1305","golang.org/x/crypto","XChaCha20-Poly1305 AEAD",                              "chacha20poly1305.NewX(key)"),
        ("go-xcrypto-argon2id",         "golang.org/x/crypto","Argon2id password hashing",                            "argon2.IDKey(password, salt, time, mem, threads, keyLen)"),
        ("go-xcrypto-argon2i",          "golang.org/x/crypto","Argon2i password hashing",                             "argon2.Key(password, salt, time, mem, threads, keyLen)"),
        ("go-xcrypto-bcrypt",           "golang.org/x/crypto","bcrypt password hashing",                              "bcrypt.GenerateFromPassword(password, cost)"),
        ("go-xcrypto-pbkdf2",           "golang.org/x/crypto","PBKDF2 key derivation (SHA-1 default)",                "pbkdf2.Key(password, salt, iter, keyLen, sha1.New)"),
        ("go-xcrypto-pbkdf2-generic",   "golang.org/x/crypto","PBKDF2 key derivation (generic hash)",                 "pbkdf2.Key(password, salt, iter, keyLen, sha256.New)"),
        ("go-xcrypto-scrypt",           "golang.org/x/crypto","scrypt key derivation",                                "scrypt.Key(password, salt, N, r, p, keyLen)"),
        ("go-xcrypto-blake2b",          "golang.org/x/crypto","BLAKE2b hash",                                         "blake2b.New256/New384/New512(key)"),
        ("go-xcrypto-blake2s",          "golang.org/x/crypto","BLAKE2s hash",                                         "blake2s.New128/New256(key)"),
        ("go-nacl-box-generatekey",     "golang.org/x/crypto","NaCl Box (Curve25519) key generation",                 "box.GenerateKey(rand)"),
        ("go-nacl-box-seal",            "golang.org/x/crypto","NaCl Box public-key encryption",                       "box.Seal(out, message, nonce, peersKey, privateKey)"),
        ("go-nacl-secretbox",           "golang.org/x/crypto","NaCl Secretbox (XSalsa20-Poly1305)",                   "secretbox.Seal(out, message, nonce, key)"),
        ("go-xcrypto-hkdf",             "golang.org/x/crypto","HKDF key derivation",                                  "hkdf.New(sha256.New, secret, salt, info)"),
        ("go-xcrypto-x25519",           "golang.org/x/crypto","X25519 Diffie-Hellman key exchange",                   "curve25519.X25519(scalar, point)"),
        ("go-xcrypto-chacha20-stream",  "golang.org/x/crypto","ChaCha20 unauthenticated stream cipher",               "chacha20.NewUnauthenticatedCipher(key, nonce)"),
        # JWT
        ("go-jwt-newwithclaims",        "golang-jwt/jwt",     "JWT token creation with signing algorithm",            "jwt.NewWithClaims(jwt.SigningMethodHS256, claims)"),
        ("go-jwt-signingmethod",        "golang-jwt/jwt",     "JWT signing method reference",                         "jwt.SigningMethodHS256 / RS256 / ES256 / EdDSA"),
        # Extras
        ("go-shamir-split",             "hashicorp/vault",    "Shamir's Secret Sharing split",                        "shamir.Split(secret, parts, threshold)"),
        ("go-totp-generate",            "pquerna/otp",        "TOTP code generation (RFC 6238)",                      "totp.GenerateCode(secret, time)"),
        ("go-hotp-generate",            "pquerna/otp",        "HOTP code generation (RFC 4226)",                      "hotp.GenerateCode(secret, counter)"),
        ("go-tink-aescmac",             "tink-crypto/tink-go","AES-CMAC via Google Tink",                             "aescmac.New(keysetHandle)"),
        ("go-certmagic-keytype",        "caddyserver/certmagic","TLS key type selection (CertMagic / ACME)",           "certmagic.KeyType = certmagic.ECDSA / RSA2048 / RSA4096"),
    ],

    # =========================================================================
    # JavaScript / TypeScript  (same rules, registered for both)
    # =========================================================================
    "JS & TypeScript": [
        # Node.js crypto
        ("js-node-crypto-createcipher",     "node:crypto",    "Cipher (deprecated — no IV)",                          "crypto.createCipher(algorithm, password)"),
        ("js-node-crypto-createcipheriv",   "node:crypto",    "Cipher with IV",                                       "crypto.createCipheriv(algorithm, key, iv)"),
        ("js-node-crypto-createdecipheriv", "node:crypto",    "Decipher with IV",                                     "crypto.createDecipheriv(algorithm, key, iv)"),
        ("js-node-crypto-createhash",       "node:crypto",    "Hash digest",                                          "crypto.createHash('sha256')"),
        ("js-node-crypto-createhmac",       "node:crypto",    "HMAC",                                                 "crypto.createHmac('sha256', key)"),
        ("js-node-crypto-generatekeypair",  "node:crypto",    "Asymmetric key pair generation",                       "crypto.generateKeyPair('rsa'/'ec'/'ed25519', ...)"),
        ("js-node-crypto-sign",             "node:crypto",    "Digital signature creation",                           "crypto.sign('sha256', data, key)"),
        ("js-node-crypto-verify",           "node:crypto",    "Digital signature verification",                       "crypto.verify('sha256', data, key, sig)"),
        ("js-node-crypto-randombytes",      "node:crypto",    "CSPRNG random bytes",                                  "crypto.randomBytes(n)"),
        ("js-node-crypto-pbkdf2",           "node:crypto",    "PBKDF2 key derivation",                                "crypto.pbkdf2(pass, salt, iter, len, digest, cb)"),
        ("js-node-crypto-scrypt",           "node:crypto",    "scrypt key derivation",                                "crypto.scrypt(pass, salt, keyLen, options, cb)"),
        ("js-node-crypto-hkdf",             "node:crypto",    "HKDF key derivation",                                  "crypto.hkdf(digest, key, salt, info, len, cb)"),
        ("js-node-crypto-dh",               "node:crypto",    "Diffie-Hellman key exchange",                          "crypto.createDiffieHellman(bits)"),
        # Web Crypto
        ("js-subtle-encrypt",               "Web Crypto API", "Symmetric/asymmetric encryption",                      "crypto.subtle.encrypt({name:'AES-GCM',...}, key, data)"),
        ("js-subtle-decrypt",               "Web Crypto API", "Symmetric/asymmetric decryption",                      "crypto.subtle.decrypt({name:'AES-GCM',...}, key, data)"),
        ("js-subtle-digest",                "Web Crypto API", "Hash digest",                                          "crypto.subtle.digest('SHA-256', data)"),
        ("js-subtle-generatekey",           "Web Crypto API", "Cryptographic key generation",                         "crypto.subtle.generateKey({name:'AES-GCM',length:256}, ...)"),
        ("js-subtle-importkey",             "Web Crypto API", "Key import from raw/JWK",                              "crypto.subtle.importKey('raw', keyData, ...)"),
        ("js-subtle-sign",                  "Web Crypto API", "Digital signature creation",                           "crypto.subtle.sign({name:'ECDSA', hash:'SHA-256'}, key, data)"),
        ("js-subtle-verify",                "Web Crypto API", "Digital signature verification",                       "crypto.subtle.verify({name:'ECDSA',...}, key, sig, data)"),
        ("js-subtle-derivekey",             "Web Crypto API", "Key derivation (PBKDF2/HKDF/ECDH)",                   "crypto.subtle.deriveKey({name:'PBKDF2',...}, baseKey, ...)"),
        ("js-subtle-derivebits",            "Web Crypto API", "Derive raw key material",                              "crypto.subtle.deriveBits({name:'ECDH',...}, key, bits)"),
        # jsonwebtoken
        ("js-jwt-sign",                     "jsonwebtoken",   "JWT token signing",                                    "jwt.sign(payload, secret, {algorithm:'HS256'})"),
        ("js-jwt-verify",                   "jsonwebtoken",   "JWT token verification",                               "jwt.verify(token, secret, {algorithms:['RS256']})"),
        ("js-jwt-none",                     "jsonwebtoken",   "JWT 'none' algorithm (no signature — vulnerable)",     "jwt.sign(payload, '', {algorithm:'none'})"),
        ("js-tls-ciphers",                  "node:tls",       "TLS server cipher string configuration",               "tls.createServer({ciphers:'ECDHE-RSA-AES256-GCM-SHA384:...'})"),
        # jose
        ("js-jose-signjwt",                 "jose",           "JWT signing with jose",                                "new SignJWT(payload).setProtectedHeader({alg:'RS256'}).sign(key)"),
        ("js-jose-encryptjwt",              "jose",           "JWT encryption with jose",                             "new EncryptJWT(payload).setProtectedHeader({alg:'RSA-OAEP',...}).encrypt(key)"),
        ("js-jose-algorithm",               "jose",           "JOSE algorithm reference",                             "jwtAlgorithm: 'RS256'/'ES256'/'PS256'/..."),
        # bcrypt
        ("js-bcrypt-hash",                  "bcrypt/bcryptjs","bcrypt password hashing",                              "bcrypt.hash(password, saltRounds)"),
        # argon2
        ("js-argon2-hash",                  "argon2",         "Argon2 password hashing",                              "argon2.hash(password, {type: argon2.argon2id})"),
        # crypto-js
        ("js-cryptojs-algo",                "crypto-js",      "Symmetric cipher (CryptoJS)",                          "CryptoJS.AES/DES/TripleDES/RC4.encrypt(msg, key)"),
        ("js-cryptojs-hmac",                "crypto-js",      "HMAC (CryptoJS)",                                      "CryptoJS.HmacSHA256(message, key)"),
        # tweetnacl
        ("js-nacl-secretbox",               "tweetnacl",      "XSalsa20-Poly1305 secretbox",                          "nacl.secretbox(msg, nonce, key)"),
        ("js-nacl-box",                     "tweetnacl",      "Curve25519 + XSalsa20-Poly1305 box",                   "nacl.box(msg, nonce, theirKey, myKey)"),
        ("js-nacl-sign",                    "tweetnacl",      "Ed25519 signature",                                    "nacl.sign(msg, secretKey)"),
        # node-forge
        ("js-forge-cipher",                 "node-forge",     "Symmetric cipher (node-forge)",                        "forge.cipher.createCipher('AES-GCM', key)"),
        ("js-forge-digest",                 "node-forge",     "Hash digest (node-forge)",                             "forge.md.sha256.create()"),
        ("js-forge-pki",                    "node-forge",     "PKI / RSA operations (node-forge)",                    "forge.pki.rsa.generateKeyPair(2048)"),
    ],

    # =========================================================================
    # Java
    # =========================================================================
    "Java": [
        # JCA
        ("jca-cipher-getInstance",          "JCA (javax.crypto)","Symmetric/asymmetric cipher",                       "Cipher.getInstance('AES/GCM/NoPadding')"),
        ("jca-messagedigest-getInstance",   "JCA (java.security)","Hash/message digest",                              "MessageDigest.getInstance('SHA-256')"),
        ("jca-signature-getInstance",       "JCA (java.security)","Digital signature algorithm",                      "Signature.getInstance('SHA256withRSA')"),
        ("jca-mac-getInstance",             "JCA (javax.crypto)","MAC (HMAC / CMAC)",                                 "Mac.getInstance('HmacSHA256')"),
        ("jca-keygenerator-getInstance",    "JCA (javax.crypto)","Symmetric key generation",                          "KeyGenerator.getInstance('AES')"),
        ("jca-keypairgenerator-getInstance","JCA (java.security)","Asymmetric key pair generation",                   "KeyPairGenerator.getInstance('RSA')"),
        ("jca-secretkeyspec",               "JCA (javax.crypto)","Secret key from raw bytes",                         "new SecretKeySpec(keyBytes, 'AES')"),
        ("jca-keyfactory-getInstance",      "JCA (java.security)","Key factory for encoding/decoding",                "KeyFactory.getInstance('RSA')"),
        ("jca-keyagreement-getInstance",    "JCA (javax.crypto)","Key agreement (DH/ECDH)",                           "KeyAgreement.getInstance('ECDH')"),
        ("jca-sslcontext-getInstance",      "JCA (javax.net.ssl)","SSL/TLS context creation",                         "SSLContext.getInstance('TLSv1.3')"),
        ("jca-secretkeyfactory-getInstance","JCA (javax.crypto)","Password-based key derivation",                     "SecretKeyFactory.getInstance('PBKDF2WithHmacSHA256')"),
        ("jca-nullcipher-new",              "JCA (javax.crypto)","NullCipher — no-op cipher (vulnerable)",            "new NullCipher()"),
        ("jca-ivparameterspec-zero",        "JCA (javax.crypto)","All-zero IV (weak — vulnerable)",                   "new IvParameterSpec(new byte[16])"),
        ("jca-securerandom-getInstance",    "JCA (java.security)","Secure random — specific algorithm",               "SecureRandom.getInstance('SHA1PRNG')"),
        ("jca-securerandom-new",            "JCA (java.security)","Secure random (default provider)",                 "new SecureRandom()"),
        ("jca-securerandom-getInstanceStrong","JCA (java.security)","Strongest secure random available",              "SecureRandom.getInstanceStrong()"),
        ("jca-keystore-getInstance",        "JCA (java.security)","KeyStore instance creation",                       "KeyStore.getInstance('PKCS12')"),
        ("jca-keystore-getKey",             "JCA (java.security)","Key retrieval from KeyStore",                      "keyStore.getKey(alias, password)"),
        ("jca-keystore-setKeyEntry",        "JCA (java.security)","Key storage into KeyStore",                        "keyStore.setKeyEntry(alias, key, password, chain)"),
        ("jca-contentSignerBuilder",        "BouncyCastle (JCA)", "X.509 certificate content signer",                 "new JcaContentSignerBuilder('SHA256WithRSAEncryption')"),
        # Commons / JWT / SSL
        ("commons-digest-utils",            "Apache Commons",    "Hash utility via Apache Commons Codec",             "DigestUtils.sha256Hex(input)"),
        ("guava-hashing",                   "Google Guava",      "Hash via Guava Hashing API",                        "Hashing.sha256().hashBytes(data)"),
        ("jjwt-sign-with",                  "jjwt",              "JWT token creation and signing",                    "Jwts.builder().signWith(key, alg).compact()"),
        ("jjwt-parse-none",                 "jjwt",              "JWT 'none' algorithm (no signature — vulnerable)",  "Jwts.parserBuilder().setAllowedClockSkewSeconds(...)"),
        ("nimbus-jws-header",               "nimbus-jose-jwt",   "JWS signed JWT (Nimbus)",                           "new JWSHeader.Builder(JWSAlgorithm.RS256).build()"),
        ("nimbus-jwe-header",               "nimbus-jose-jwt",   "JWE encrypted JWT (Nimbus)",                        "new JWEHeader.Builder(JWEAlgorithm.RSA_OAEP_256, ...).build()"),
        ("ssl-socket-enabled-protocols",    "javax.net.ssl",     "Enabled TLS protocol versions",                    "socket.setEnabledProtocols(new String[]{'TLSv1.2'})"),
        ("ssl-socket-enabled-suites",       "javax.net.ssl",     "Enabled TLS cipher suites",                        "socket.setEnabledCipherSuites(new String[]{...})"),
        # Spring Security
        ("spring-md5-password-encoder",     "Spring Security",   "MD5 password encoder (deprecated — vulnerable)",   "new Md5PasswordEncoder()"),
        ("spring-sha-password-encoder",     "Spring Security",   "SHA password encoder (deprecated — vulnerable)",   "new ShaPasswordEncoder()"),
        ("spring-bcrypt-password-encoder",  "Spring Security",   "bcrypt password encoder",                          "new BCryptPasswordEncoder()"),
        ("spring-argon2-password-encoder",  "Spring Security",   "Argon2 password encoder",                          "new Argon2PasswordEncoder(...)"),
        ("spring-pbkdf2-password-encoder",  "Spring Security",   "PBKDF2 password encoder",                          "new Pbkdf2PasswordEncoder(secret, iter, hashWidth)"),
        ("spring-scrypt-password-encoder",  "Spring Security",   "scrypt password encoder",                          "new SCryptPasswordEncoder(...)"),
        # BouncyCastle
        ("bc-block-cipher-engine",          "BouncyCastle",      "Block cipher engine",                              "new AESEngine/DESEngine/BlowfishEngine/..."),
        ("bc-stream-cipher-engine",         "BouncyCastle",      "Stream cipher engine",                             "new RC4Engine/ChaCha7539Engine/..."),
        ("bc-aead-cipher",                  "BouncyCastle",      "AEAD cipher (GCM / CCM / OCB / EAX)",              "new GCMBlockCipher/CCMBlockCipher/OCBBlockCipher(...)"),
        ("bc-digest",                       "BouncyCastle",      "Hash digest",                                      "new SHA256Digest/SHA3Digest/Blake2bDigest/...()"),
        ("bc-mac",                          "BouncyCastle",      "MAC (HMAC / GMAC / CMAC / Poly1305)",              "new HMac/GMac/CMac/Poly1305(cipher)"),
        ("bc-keygen",                       "BouncyCastle",      "Symmetric key generation",                         "new CipherKeyGenerator().init(new KeyGenerationParameters(...))"),
        ("bc-signer",                       "BouncyCastle",      "Asymmetric signer (RSA / ECDSA / Ed25519)",        "new RSADigestSigner/ECDSASigner/Ed25519Signer()"),
        ("bc-keypairgen",                   "BouncyCastle",      "Asymmetric key pair generation",                   "new RSAKeyPairGenerator/ECKeyPairGenerator/...().generateKeyPair()"),
        ("bc-keyagreement",                 "BouncyCastle",      "Key agreement (ECDH / DH)",                        "new ECDHBasicAgreement/DHBasicAgreement()"),
        ("bc-pbe-paramgen",                 "BouncyCastle",      "Password-Based Encryption parameter generation",   "new PKCS5S2ParametersGenerator/PKCS12ParametersGenerator(...)"),
    ],

    # =========================================================================
    # PHP
    # =========================================================================
    "PHP": [
        # openssl_*
        ("php-openssl-encrypt",         "openssl",        "Symmetric encryption",                                     "openssl_encrypt($data, 'AES-256-GCM', $key, ...)"),
        ("php-openssl-decrypt",         "openssl",        "Symmetric decryption",                                     "openssl_decrypt($data, 'AES-256-GCM', $key, ...)"),
        ("php-openssl-sign",            "openssl",        "Asymmetric digital signature",                             "openssl_sign($data, $sig, $key, OPENSSL_ALGO_SHA256)"),
        ("php-openssl-verify",          "openssl",        "Asymmetric signature verification",                        "openssl_verify($data, $sig, $key, OPENSSL_ALGO_SHA256)"),
        ("php-openssl-pkey-new",        "openssl",        "Asymmetric key pair generation",                           "openssl_pkey_new(['private_key_type'=>OPENSSL_KEYTYPE_RSA])"),
        ("php-openssl-digest",          "openssl",        "Hash/message digest",                                      "openssl_digest($data, 'sha256')"),
        ("php-openssl-random",          "openssl",        "CSPRNG random bytes",                                      "openssl_random_pseudo_bytes($n)"),
        ("php-openssl-public-encrypt",  "openssl",        "RSA public-key encryption",                                "openssl_public_encrypt($data, $enc, $key)"),
        ("php-openssl-private-decrypt", "openssl",        "RSA private-key decryption",                               "openssl_private_decrypt($enc, $plain, $key)"),
        ("php-openssl-private-encrypt", "openssl",        "RSA private-key encryption (signing)",                     "openssl_private_encrypt($data, $enc, $key)"),
        ("php-openssl-public-decrypt",  "openssl",        "RSA public-key decryption (verify)",                       "openssl_public_decrypt($enc, $plain, $key)"),
        ("php-openssl-seal",            "openssl",        "Envelope encryption (seal)",                               "openssl_seal($data, $sealed, $envKeys, $pubKeys)"),
        ("php-openssl-open",            "openssl",        "Envelope decryption (open)",                               "openssl_open($sealed, $plain, $envKey, $privKey)"),
        ("php-openssl-pkcs7-sign",      "openssl",        "PKCS#7 / S/MIME signing",                                  "openssl_pkcs7_sign($infile, $outfile, $cert, $key, [])"),
        ("php-openssl-pkcs7-verify",    "openssl",        "PKCS#7 / S/MIME verification",                             "openssl_pkcs7_verify($infile, 0, $signers)"),
        ("php-openssl-csr-new",         "openssl",        "CSR / X.509 certificate generation",                       "openssl_csr_new($dn, $privKey)"),
        ("php-openssl-encrypt-dynamic", "openssl",        "Dynamic cipher string in openssl_encrypt",                 "openssl_encrypt($d, $cipherVar, $k)"),
        ("php-openssl-decrypt-dynamic", "openssl",        "Dynamic cipher string in openssl_decrypt",                 "openssl_decrypt($d, $cipherVar, $k)"),
        # hash / HMAC / PBKDF2
        ("php-hash",                    "hash",           "Hash / HMAC / hash_equals",                                "hash('sha256', $data) / hash_equals($a,$b)"),
        ("php-hash-hmac",               "hash",           "HMAC message authentication code",                        "hash_hmac('sha256', $data, $key)"),
        ("php-hash-pbkdf2",             "hash",           "PBKDF2 key derivation",                                    "hash_pbkdf2('sha256', $pass, $salt, $iter, $len)"),
        ("php-hash-init",               "hash",           "Incremental hash context",                                 "hash_init('sha256', HASH_HMAC, $key)"),
        ("php-md5",                     "hash",           "MD5 hash (built-in)",                                      "md5($data) / md5_file($path)"),
        ("php-sha1",                    "hash",           "SHA-1 hash (built-in)",                                    "sha1($data) / sha1_file($path)"),
        # password_hash
        ("php-password-hash",           "password",       "Password hashing (bcrypt / argon2id / argon2i)",           "password_hash($pass, PASSWORD_BCRYPT / PASSWORD_ARGON2ID)"),
        ("php-password-hash-dynamic",   "password",       "Dynamic algorithm in password_hash",                       "password_hash($pass, $algo)"),
        ("php-password-verify",         "password",       "Password verification",                                    "password_verify($plain, $hash)"),
        # sodium
        ("php-sodium-aead",             "sodium",         "AEAD encryption (sodium)",                                 "sodium_crypto_aead_chacha20poly1305_encrypt/aes256gcm_encrypt(...)"),
        ("php-sodium-secretbox",        "sodium",         "XSalsa20-Poly1305 secretbox",                              "sodium_crypto_secretbox($msg, $nonce, $key)"),
        ("php-sodium-box",              "sodium",         "Curve25519 + XSalsa20-Poly1305 box",                       "sodium_crypto_box($msg, $nonce, $kp)"),
        ("php-sodium-sign",             "sodium",         "Ed25519 signing",                                          "sodium_crypto_sign_detached($msg, $secretKey)"),
        ("php-sodium-hash",             "sodium",         "BLAKE2b hash (sodium)",                                    "sodium_crypto_generichash($msg, $key)"),
        ("php-sodium-pwhash",           "sodium",         "Argon2id password hashing (sodium)",                       "sodium_crypto_pwhash($len, $pass, $salt, ...)"),
        ("php-sodium-kdf",              "sodium",         "HKDF-like key derivation (sodium)",                        "sodium_crypto_kdf_derive_from_key($len, $id, $ctx, $key)"),
        ("php-sodium-scalarmult",       "sodium",         "X25519 Diffie-Hellman (sodium)",                           "sodium_crypto_scalarmult($n, $p)"),
        ("php-sodium-auth",             "sodium",         "HMAC-SHA-512/256 authentication (sodium)",                 "sodium_crypto_auth($msg, $key)"),
        ("php-sodium-shorthash",        "sodium",         "SipHash-2-4 short hash (sodium)",                          "sodium_crypto_shorthash($msg, $key)"),
        # mcrypt
        ("php-mcrypt-encrypt",          "mcrypt",         "mcrypt encryption (deprecated PHP < 7.1)",                 "mcrypt_encrypt(MCRYPT_RIJNDAEL_128, $key, $data, ...)"),
        # TLS
        ("php-stream-ssl-ciphers",      "stream_context", "TLS cipher string in stream context",                      "stream_context_create(['ssl'=>['ciphers'=>'ECDHE+AES...']])"),
        # Random
        ("php-random-bytes",            "random",         "CSPRNG random bytes (PHP 7+)",                             "random_bytes($n)"),
        ("php-random-int",              "random",         "CSPRNG random integer (PHP 7+)",                           "random_int($min, $max)"),
        ("php-crypt",                   "crypt",          "crypt() DES / MD5 / bcrypt password hash",                 "crypt($pass, $salt)"),
        # Firebase JWT
        ("php-firebase-jwt-encode",     "firebase/php-jwt","JWT encode (Firebase)",                                   "JWT::encode($payload, $key, 'RS256')"),
        ("php-firebase-jwt-decode",     "firebase/php-jwt","JWT decode (Firebase)",                                   "JWT::decode($token, new Key($key, 'RS256'))"),
        # phpseclib
        ("php-phpseclib-aes",           "phpseclib",      "AES cipher (phpseclib class API)",                         "new AES('cbc') / $aes->setKey($k) / $aes->encrypt($d)"),
        ("php-phpseclib-tripledes",     "phpseclib",      "TripleDES / 3DES cipher (phpseclib)",                      "new TripleDES('cbc')"),
        ("php-phpseclib-des",           "phpseclib",      "DES cipher (phpseclib)",                                   "new DES('ecb')"),
        ("php-phpseclib-rijndael",      "phpseclib",      "Rijndael cipher (phpseclib)",                              "new Rijndael('cbc')"),
        ("php-phpseclib-blowfish",      "phpseclib",      "Blowfish cipher (phpseclib)",                              "new Blowfish('cbc')"),
        ("php-phpseclib-twofish",       "phpseclib",      "Twofish cipher (phpseclib)",                               "new Twofish('cbc')"),
        ("php-phpseclib-rc4",           "phpseclib",      "RC4 stream cipher (phpseclib)",                            "new RC4()"),
        ("php-phpseclib-chacha20",      "phpseclib",      "ChaCha20 stream cipher (phpseclib)",                       "new ChaCha20()"),
        ("php-phpseclib-salsa20",       "phpseclib",      "Salsa20 stream cipher (phpseclib)",                        "new Salsa20()"),
        ("php-phpseclib-rsa-createkey", "phpseclib",      "RSA key generation (phpseclib)",                           "RSA::createKey($bits)"),
        ("php-phpseclib-ec-createkey",  "phpseclib",      "EC key generation (phpseclib)",                            "EC::createKey('P-256' / 'Ed25519' / ...)"),
        ("php-phpseclib-dsa-createkey", "phpseclib",      "DSA key generation (phpseclib)",                           "DSA::createKey($bits)"),
        ("php-phpseclib-dh-createkey",  "phpseclib",      "DH key generation (phpseclib)",                            "DH::createKey($primeBits)"),
        ("php-phpseclib-dh-computesecret","phpseclib",    "DH shared secret computation (phpseclib)",                 "DH::computeSecret($privateKey, $publicKey)"),
        ("php-phpseclib-hash",          "phpseclib",      "Hash algorithm (phpseclib)",                               "new Hash('sha256') / $h->hash($data)"),
        ("php-phpseclib-random",        "phpseclib",      "CSPRNG random bytes (phpseclib)",                          "Random::string($n)"),
    ],

    # =========================================================================
    # C#
    # =========================================================================
    "C#": [
        # Symmetric
        ("cs-aes-create",               "System.Security.Cryptography","AES cipher (default)",                        "Aes.Create()"),
        ("cs-aesgcm-new",               "System.Security.Cryptography","AES-GCM AEAD",                                "new AesGcm(key)"),
        ("cs-des-create",               "System.Security.Cryptography","DES cipher (deprecated — vulnerable)",        "DES.Create()"),
        ("cs-tripledes-create",         "System.Security.Cryptography","TripleDES cipher (deprecated)",               "TripleDES.Create()"),
        ("cs-rc2-create",               "System.Security.Cryptography","RC2 cipher (deprecated — vulnerable)",        "RC2.Create()"),
        ("cs-aes-csp",                  "System.Security.Cryptography","AES CSP implementation",                      "new AesCryptoServiceProvider()"),
        ("cs-aes-cng",                  "System.Security.Cryptography","AES CNG implementation",                      "new AesCng()"),
        ("cs-rijndael-managed",         "System.Security.Cryptography","RijndaelManaged cipher (deprecated)",         "new RijndaelManaged()"),
        ("cs-ciphermode",               "System.Security.Cryptography","Cipher mode assignment",                      "algorithm.Mode = CipherMode.CBC/ECB/GCM/..."),
        ("cs-aesccm-new",               "System.Security.Cryptography","AES-CCM AEAD",                                "new AesCcm(key)"),
        ("cs-chacha20poly1305-new",     "System.Security.Cryptography","ChaCha20-Poly1305 AEAD",                      "new ChaCha20Poly1305(key)"),
        ("cs-create-encryptor",         "System.Security.Cryptography","Create ICryptoTransform encryptor",           "algorithm.CreateEncryptor()"),
        ("cs-create-decryptor",         "System.Security.Cryptography","Create ICryptoTransform decryptor",           "algorithm.CreateDecryptor()"),
        # Hash
        ("cs-sha256-create",            "System.Security.Cryptography","SHA-256 hash",                                "SHA256.Create()"),
        ("cs-sha512-create",            "System.Security.Cryptography","SHA-512 hash",                                "SHA512.Create()"),
        ("cs-sha384-create",            "System.Security.Cryptography","SHA-384 hash",                                "SHA384.Create()"),
        ("cs-sha1-create",              "System.Security.Cryptography","SHA-1 hash (deprecated — vulnerable)",        "SHA1.Create()"),
        ("cs-md5-create",               "System.Security.Cryptography","MD5 hash (deprecated — vulnerable)",          "MD5.Create()"),
        ("cs-sha3-create",              "System.Security.Cryptography","SHA-3 hash (SHA3-256/384/512)",               "SHA3_256.Create() / SHA3_384.Create() / SHA3_512.Create()"),
        # MAC
        ("cs-hmac-new",                 "System.Security.Cryptography","HMAC message authentication code",            "new HMACSHA256(key) / HMACSHA512 / HMACSHA384 / HMACMD5"),
        # Asymmetric
        ("cs-rsa-create",               "System.Security.Cryptography","RSA key creation",                            "RSA.Create()"),
        ("cs-rsa-create-size",          "System.Security.Cryptography","RSA key creation with key size",              "RSA.Create(2048)"),
        ("cs-ecdsa-create",             "System.Security.Cryptography","ECDSA key creation",                          "ECDsa.Create()"),
        ("cs-ecdh-create",              "System.Security.Cryptography","ECDH key agreement",                          "ECDiffieHellman.Create()"),
        ("cs-dsa-create",               "System.Security.Cryptography","DSA key creation",                            "DSA.Create()"),
        # KDF
        ("cs-pbkdf2-rfc2898",           "System.Security.Cryptography","PBKDF2 key derivation",                       "new Rfc2898DeriveBytes(password, salt, iterations)"),
        ("cs-keyderivation-pbkdf2",     "Microsoft.AspNetCore.Cryptography.KeyDerivation","ASP.NET Core PBKDF2","KeyDerivation.Pbkdf2(password, salt, prf, iter, len)"),
        ("cs-keyderivation-prf",        "Microsoft.AspNetCore.Cryptography.KeyDerivation","PBKDF2 PRF enum",          "KeyDerivationPrf.HMACSHA256 / HMACSHA512"),
        # CSPRNG
        ("cs-rng",                      "System.Security.Cryptography","CSPRNG random bytes",                         "RandomNumberGenerator.Fill(buf) / .GetBytes(buf)"),
        # TLS
        ("cs-sslstream-new",            "System.Net.Security",         "TLS SslStream creation",                      "new SslStream(inner, false, certCallback)"),
        ("cs-sslstream-authenticate",   "System.Net.Security",         "TLS authentication",                          "stream.AuthenticateAsClient/Server(host, ...)"),
        # JWT
        ("cs-jwt-securitytoken",        "System.IdentityModel.Tokens.Jwt","JWT token creation",                       "new JwtSecurityToken(issuer, aud, claims, ...)"),
        ("cs-security-algorithms",      "System.IdentityModel.Tokens.Jwt","JWT signing/encryption algorithm constants","SecurityAlgorithms.HmacSha256/RsaSha256/Aes256CbcHmacSha512"),
        ("cs-json-web-token-handler",   "System.IdentityModel.Tokens.Jwt","JWT handler (create / validate tokens)",   "new JsonWebTokenHandler()"),
        # X.509
        ("cs-x509certificate2",         "System.Security.Cryptography.X509Certificates","X.509 certificate loading","new X509Certificate2(path / bytes / thumbprint)"),
        # DataProtection enums
        ("cs-dp-encryption-algo",       "Microsoft.AspNetCore.DataProtection","DataProtection encryption algorithm enum","EncryptionAlgorithm.AES_256_CBC / AES_128_GCM / AES_256_GCM"),
        ("cs-dp-validation-algo",       "Microsoft.AspNetCore.DataProtection","DataProtection validation algorithm enum","ValidationAlgorithm.HMACSHA256 / HMACSHA512"),
        # Post-quantum
        ("cs-mldsa",                    "System.Security.Cryptography","ML-DSA (FIPS 204) key generation",            "MLDsa.GenerateKey(MLDsaAlgorithm.MLDsa44/65/87)"),
        ("cs-slhdsa",                   "System.Security.Cryptography","SLH-DSA (FIPS 205) key generation",           "SlhDsa.GenerateKey(SlhDsaAlgorithm.SLHDsaSha2_128s/...)"),
        ("cs-composite-mldsa",          "System.Security.Cryptography","Composite ML-DSA key generation",             "CompositeMLDsa.GenerateKey(CompositeMLDsaAlgorithm.MLDSA44WithRSA2048Pss/...)"),
        # BouncyCastle .NET
        ("cs-bc-block-cipher",          "BouncyCastle .NET",           "BouncyCastle block cipher",                   "new AesEngine/DesEngine/TwofishEngine/...()"),
        ("cs-bc-digest",                "BouncyCastle .NET",           "BouncyCastle hash digest",                    "new Sha256Digest/Sha3Digest/Blake2bDigest/...()"),
        ("cs-bc-mac",                   "BouncyCastle .NET",           "BouncyCastle MAC",                            "new HMac/GMac/CMac/Poly1305(cipher)"),
        ("cs-bc-signer",                "BouncyCastle .NET",           "BouncyCastle asymmetric signer",              "new RsaDigestSigner/ECDsaSigner/Ed25519Signer()"),
        ("cs-bc-keypairgen",            "BouncyCastle .NET",           "BouncyCastle key pair generation",            "new RsaKeyPairGenerator/ECKeyPairGenerator/...()"),
        ("cs-bc-pbe",                   "BouncyCastle .NET",           "BouncyCastle password-based encryption",      "new Pkcs5S2ParametersGenerator/Pkcs12ParametersGenerator(digest)"),
    ],

    # =========================================================================
    # Ruby
    # =========================================================================
    "Ruby": [
        ("ruby-openssl-cipher-new",     "openssl",        "Generic cipher (OpenSSL::Cipher.new)",                     "OpenSSL::Cipher.new('AES-256-GCM')"),
        ("ruby-openssl-cipher-aes",     "openssl",        "AES shorthand cipher",                                     "OpenSSL::Cipher::AES.new(256, :GCM)"),
        ("ruby-openssl-digest-new",     "openssl",        "Generic digest (OpenSSL::Digest.new)",                     "OpenSSL::Digest.new('SHA256')"),
        ("ruby-openssl-digest-class",   "openssl",        "Named digest class shorthand",                             "OpenSSL::Digest::SHA256.new"),
        ("ruby-openssl-hmac",           "openssl",        "HMAC (OpenSSL::HMAC)",                                     "OpenSSL::HMAC.digest('SHA256', key, data)"),
        ("ruby-openssl-pkey-rsa",       "openssl",        "RSA key generation",                                       "OpenSSL::PKey::RSA.new(2048)"),
        ("ruby-openssl-pkey-ec",        "openssl",        "EC key generation",                                        "OpenSSL::PKey::EC.new('prime256v1')"),
        ("ruby-openssl-pkey-dsa",       "openssl",        "DSA key generation",                                       "OpenSSL::PKey::DSA.new(2048)"),
        ("ruby-openssl-pkey-dh",        "openssl",        "DH key generation",                                        "OpenSSL::PKey::DH.new(2048)"),
        ("ruby-openssl-pkey-generate",  "openssl",        "Generic key generation",                                   "OpenSSL::PKey.generate_key('RSA', rsa_keygen_bits: 2048)"),
        ("ruby-openssl-pkey-sign",      "openssl",        "PKey signing",                                             "key.sign(OpenSSL::Digest::SHA256.new, data)"),
        ("ruby-openssl-pkey-verify",    "openssl",        "PKey signature verification",                              "key.verify(OpenSSL::Digest::SHA256.new, sig, data)"),
        ("ruby-openssl-pkey-derive",    "openssl",        "ECDH / DH key derivation",                                 "key.derive(peer_key)"),
        ("ruby-openssl-kdf",            "openssl",        "Key derivation (PBKDF2 / HKDF / scrypt)",                  "OpenSSL::KDF.pbkdf2_hmac(pass, salt:, iter:, length:, hash:)"),
        ("ruby-openssl-ssl-context",    "openssl",        "SSL/TLS context",                                          "OpenSSL::SSL::SSLContext.new"),
        ("ruby-openssl-ssl-socket",     "openssl",        "SSL/TLS socket",                                           "OpenSSL::SSL::SSLSocket.new(socket, ctx)"),
        ("ruby-openssl-ssl-ciphers",    "openssl",        "TLS cipher string configuration",                          "ctx.ciphers = 'ECDHE+AESGCM:ECDHE+CHACHA20:...'"),
        ("ruby-openssl-ssl-version",    "openssl",        "TLS version selection",                                    "ctx.ssl_version = :TLSv1_2 / :TLSv1_3"),
        ("ruby-openssl-pkcs5-pbkdf2",   "openssl",        "PBKDF2-HMAC key derivation",                               "OpenSSL::PKCS5.pbkdf2_hmac(pass, salt, iter, len, digest)"),
        ("ruby-openssl-random",         "openssl",        "CSPRNG random bytes",                                      "OpenSSL::Random.random_bytes(n)"),
        ("ruby-digest-class",           "digest",         "Digest class reference",                                   "Digest::SHA256 / Digest::MD5 / Digest::SHA512"),
        ("ruby-digest-hexdigest",       "digest",         "Hex digest shorthand",                                     "Digest::SHA256.hexdigest(data)"),
        ("ruby-bcrypt-create",          "bcrypt",         "bcrypt password hash (BCrypt::Password.create)",           "BCrypt::Password.create(password, cost: 12)"),
        ("ruby-bcrypt-password-new",    "bcrypt",         "bcrypt password verification",                             "BCrypt::Password.new(hash)"),
        ("ruby-bcrypt-engine-hash",     "bcrypt",         "bcrypt raw engine hash",                                   "BCrypt::Engine.hash_secret(secret, salt)"),
        ("ruby-jwt-encode",             "ruby-jwt",       "JWT encode (token creation)",                              "JWT.encode(payload, key, 'RS256')"),
        ("ruby-jwt-decode",             "ruby-jwt",       "JWT decode (token verification)",                          "JWT.decode(token, pub_key, true, {algorithm:'RS256'})"),
        ("ruby-securerandom",           "securerandom",   "SecureRandom CSPRNG",                                      "SecureRandom.random_bytes(n) / SecureRandom.hex(n)"),
    ],

    # =========================================================================
    # Rust
    # =========================================================================
    "Rust": [
        # ring
        ("rust-ring-aead",              "ring",           "AEAD (AES-GCM / ChaCha20-Poly1305)",                       "ring::aead::SealingKey / OpeningKey / UnboundKey::new(CHACHA20_POLY1305, ...)"),
        ("rust-ring-digest",            "ring",           "Hash digest",                                              "ring::digest::digest(ring::digest::SHA256, data)"),
        ("rust-ring-hmac",              "ring",           "HMAC",                                                     "ring::hmac::Key::new(ring::hmac::HMAC_SHA256, key)"),
        ("rust-ring-signature",         "ring",           "Digital signature (RSA / ECDSA / Ed25519)",                "ring::signature::Ed25519KeyPair / EcdsaKeyPair"),
        ("rust-ring-pbkdf2",            "ring",           "PBKDF2 key derivation",                                    "ring::pbkdf2::derive(ring::pbkdf2::PBKDF2_HMAC_SHA256, ...)"),
        ("rust-ring-random",            "ring",           "CSPRNG random bytes",                                      "ring::rand::SystemRandom::new().fill(&mut buf)"),
        # RustCrypto AEAD
        ("rust-aes-gcm",                "aes-gcm",        "AES-GCM AEAD (RustCrypto)",                               "Aes256Gcm::new(key) / Aes128Gcm::new(key)"),
        ("rust-chacha20poly1305",        "chacha20poly1305","ChaCha20-Poly1305 / XChaCha20-Poly1305 AEAD",            "ChaCha20Poly1305::new(key) / XChaCha20Poly1305::new(key)"),
        ("rust-aes-siv",                "aes-siv",        "AES-SIV deterministic AEAD",                              "Aes256SivAead::new(key)"),
        # Symmetric
        ("rust-aes-cbc",                "aes/cbc",        "AES-CBC block cipher",                                     "cbc::Encryptor::<aes::Aes256>::new(key, iv)"),
        ("rust-aes-ctr",                "aes/ctr",        "AES-CTR stream cipher",                                    "ctr::Ctr128BE::<aes::Aes256>::new(key, nonce)"),
        # Hash
        ("rust-sha2",                   "sha2",           "SHA-2 hash (SHA-256 / SHA-512)",                           "Sha256::new() / Sha512::digest(data)"),
        ("rust-sha3",                   "sha3",           "SHA-3 hash / Keccak",                                      "Sha3_256::new() / Keccak256::digest(data)"),
        ("rust-md5",                    "md5",            "MD5 hash (deprecated — vulnerable)",                       "md5::compute(data)"),
        ("rust-blake2",                 "blake2",         "BLAKE2b / BLAKE2s hash",                                   "Blake2b512::new() / Blake2s256::digest(data)"),
        ("rust-blake3",                 "blake3",         "BLAKE3 hash",                                              "blake3::hash(data) / blake3::Hasher::new()"),
        # HMAC
        ("rust-hmac",                   "hmac",           "HMAC (RustCrypto)",                                        "Hmac::<Sha256>::new_from_slice(key)"),
        # RSA
        ("rust-rsa-privatekey",         "rsa",            "RSA private key generation",                               "RsaPrivateKey::new(&mut rng, bits)"),
        ("rust-rsa-sign",               "rsa",            "RSA signing / OAEP encryption",                            "key.sign(PaddingScheme::new_pkcs1v15_sign::<Sha256>(), ...)"),
        # Ed25519
        ("rust-ed25519-signingkey",     "ed25519-dalek",  "Ed25519 signing key",                                      "SigningKey::generate(&mut rng)"),
        ("rust-ed25519-verifyingkey",   "ed25519-dalek",  "Ed25519 verifying key",                                    "VerifyingKey::from(&signing_key)"),
        # X25519
        ("rust-x25519",                 "x25519-dalek",   "X25519 Diffie-Hellman key exchange",                       "EphemeralSecret::random_from_rng(rng) / StaticSecret::random_from_rng(rng)"),
        # ECDSA
        ("rust-ecdsa",                  "p256/p384",      "ECDSA signing (P-256 / P-384)",                            "p256::ecdsa::SigningKey::random(&mut rng)"),
        # Password hashing
        ("rust-argon2",                 "argon2",         "Argon2 password hashing",                                  "Argon2::default().hash_password(password, &salt)"),
        ("rust-bcrypt",                 "bcrypt",         "bcrypt password hashing",                                  "bcrypt::hash(password, bcrypt::DEFAULT_COST)"),
        ("rust-scrypt",                 "scrypt",         "scrypt key derivation",                                     "scrypt::scrypt(password, salt, &params, output)"),
        ("rust-pbkdf2",                 "pbkdf2",         "PBKDF2 key derivation (standalone crate)",                 "pbkdf2::pbkdf2_hmac::<Sha256>(password, salt, rounds, out)"),
        # rustls
        ("rust-rustls-client",          "rustls",         "TLS client config (builder)",                              "rustls::ClientConfig::builder().with_root_certificates(...)"),
        ("rust-rustls-server",          "rustls",         "TLS server config (builder)",                              "rustls::ServerConfig::builder().with_no_client_auth()"),
        ("rust-rustls-client-with-provider","rustls",     "TLS client config with custom crypto provider",            "rustls::ClientConfig::builder_with_provider(provider)"),
        ("rust-rustls-server-with-provider","rustls",     "TLS server config with custom crypto provider",            "rustls::ServerConfig::builder_with_provider(provider)"),
        ("rust-rustls-ciphersuite",     "rustls",         "TLS cipher suite constant",                                "rustls::cipher_suite::TLS13_AES_256_GCM_SHA384"),
        ("rust-rustls-sigscheme",       "rustls",         "TLS signature scheme constant",                            "rustls::SignatureScheme::ECDSA_NISTP256_SHA256"),
        ("rust-sm4",                    "rustls / ring",  "SM4 block cipher (TLS SM4-GCM / SM4-CCM)",                 "rustls::cipher_suite::TLS13_SM4_GCM_SM3"),
        ("rust-hkdf",                   "hkdf",           "HKDF key derivation",                                      "Hkdf::<Sha256>::new(salt, ikm).expand(info, okm)"),
        # JWT
        ("rust-jwt-encode",             "jsonwebtoken",   "JWT encode (token creation)",                              "encode(&Header::new(Algorithm::RS256), &claims, &key)"),
        ("rust-jwt-decode",             "jsonwebtoken",   "JWT decode (token verification)",                          "decode::<Claims>(&token, &key, &Validation::new(Algorithm::RS256))"),
        ("rust-jwt-algorithm",          "jsonwebtoken",   "JWT algorithm reference",                                  "Algorithm::HS256 / RS256 / ES256 / EdDSA / PS256"),
        # SQLCipher
        ("rust-sqlcipher-pragma-key",   "sqlcipher (libsqlite3-sys)","SQLCipher AES-256 database encryption key setup","format!(\"PRAGMA key = \\\"{}\\\"\\nPRAGMA rekey = ...\")"),
    ],

    # =========================================================================
    # Dart / Flutter
    # =========================================================================
    "Dart-Flutter": [
        # package:crypto
        ("dart-crypto-digest",          "package:crypto",       "Hash digest (MD5/SHA1/SHA256/SHA512)",              "md5.convert(bytes) / sha256.convert(bytes)"),
        ("dart-crypto-hmac",            "package:crypto",       "HMAC",                                              "Hmac(sha256, key).convert(bytes)"),
        # package:cryptography
        ("dart-cryptography-aes",       "package:cryptography", "AES-CBC / AES-CTR / AES-GCM",                      "AesCbc() / AesCtr() / AesGcm()"),
        ("dart-cryptography-chacha20",  "package:cryptography", "ChaCha20-Poly1305 / XChaCha20-Poly1305",           "Chacha20.poly1305Aead() / XChacha20.poly1305Aead()"),
        ("dart-cryptography-hash",      "package:cryptography", "Hash algorithm (Sha256 / Blake2b / Md5 / …)",      "Sha256() / Sha512() / Blake2b() / Blake2s()"),
        ("dart-cryptography-hmac",      "package:cryptography", "HMAC via named constructor",                       "Hmac.sha256() / Hmac.sha512()"),
        ("dart-cryptography-ed25519",   "package:cryptography", "Ed25519 signing key pair",                         "Ed25519()"),
        ("dart-cryptography-x25519",    "package:cryptography", "X25519 key agreement",                             "X25519()"),
        ("dart-cryptography-ecdh",      "package:cryptography", "ECDH key agreement (P-256/P-384/P-521)",           "Ecdh.p256() / Ecdh.p384() / Ecdh.p521()"),
        ("dart-cryptography-ecdsa",     "package:cryptography", "ECDSA signing (P-256/P-384/P-521)",                "Ecdsa.p256(Sha256()) / Ecdsa.p384(Sha384())"),
        ("dart-cryptography-rsa",       "package:cryptography", "RSA-PSS / RSA-PKCS#1 v1.5 signing",               "RsaPss(...) / RsaSsaPkcs1v15(...)"),
        ("dart-cryptography-kdf",       "package:cryptography", "Key derivation (PBKDF2 / HKDF / Argon2id)",        "Pbkdf2(...) / Hkdf(...) / Argon2id(...)"),
        ("dart-cryptography-secretbox", "package:cryptography", "SecretBox AEAD",                                   "SecretBox(ciphertext, nonce: nonce, mac: mac)"),
        # package:pointycastle
        ("dart-pointycastle-blockcipher","package:pointycastle","Block cipher (BouncyCastle port)",                  "BlockCipher('AES/CBC') / BlockCipher('DES/ECB')"),
        ("dart-pointycastle-paddedblockcipher","package:pointycastle","Padded block cipher",                        "PaddedBlockCipher('AES/CBC/PKCS7')"),
        ("dart-pointycastle-streamcipher","package:pointycastle","Stream cipher (RC4, ChaCha20…)",                  "StreamCipher('RC4')"),
        ("dart-pointycastle-digest",    "package:pointycastle", "Hash digest (BouncyCastle port)",                  "Digest('SHA-256') / Digest('MD5')"),
        ("dart-pointycastle-mac",       "package:pointycastle", "MAC (HMAC / CMAC)",                                "Mac('SHA-256/HMAC') / Mac('AES/CMAC')"),
        ("dart-pointycastle-keygen",    "package:pointycastle", "Symmetric key generation",                         "KeyGenerator('AES')"),
        ("dart-pointycastle-signer",    "package:pointycastle", "Asymmetric signer",                                "Signer('RSA') / Signer('ECDSA')"),
        ("dart-pointycastle-asymmetric-keypairgen","package:pointycastle","Asymmetric key pair generation",         "AsymmetricKeyPairGenerator('RSA')"),
        ("dart-pointycastle-kdf",       "package:pointycastle", "Key derivation (PBKDF2/scrypt/HKDF)",              "PBKDF2KeyDerivator() / ScryptParameters() / HkdfParameters()"),
        # package:encrypt
        ("dart-encrypt-aes",            "package:encrypt",      "AES encryption (high-level wrapper)",              "AES(key, mode: AESMode.cbc)"),
        ("dart-encrypt-rsa",            "package:encrypt",      "RSA encryption (high-level wrapper)",              "RSA(publicKey: key)"),
        ("dart-encrypt-fernet",         "package:encrypt",      "Fernet (AES-128-CBC + HMAC-SHA256)",               "Fernet(key)"),
        # dart:io TLS
        ("dart-tls-securesocket",       "dart:io",              "TLS secure socket connection",                     "SecureSocket.connect/RawSecureSocket.connect/bind(...)"),
        ("dart-tls-securitycontext",    "dart:io",              "TLS security context",                             "SecurityContext() / SecurityContext.defaultContext"),
        # CSPRNG
        ("dart-random-secure",          "dart:math",            "CSPRNG (cryptographically secure random)",         "Random.secure()"),
        # JWT
        ("dart-jwt-sign",               "dart_jsonwebtoken/jose","JWT signing algorithm reference",                  "JWTAlgorithm.HS256/RS256/ES256/EdDSA/none"),
        # isar (SQLCipher)
        ("dart-isar-encryption-key",    "isar (SQLCipher)",     "SQLCipher AES-256 database encryption",            "Isar.open(encryptionKey: key) / Isar.openAsync(encryptionKey: key)"),
        ("dart-isar-change-encryption-key","isar (SQLCipher)",  "SQLCipher database key rotation",                  "isar.changeEncryptionKey(newKey)"),
    ],
}

# ---------------------------------------------------------------------------
# Styling helpers
# ---------------------------------------------------------------------------

HEADER_FILL   = PatternFill("solid", fgColor="1F3864")   # dark navy
ALT_FILL      = PatternFill("solid", fgColor="EBF0FA")   # light blue-gray
WHITE_FILL    = PatternFill("solid", fgColor="FFFFFF")
HEADER_FONT   = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
BODY_FONT     = Font(name="Calibri", size=10)
TITLE_FONT    = Font(bold=True, name="Calibri", size=13)
THIN_BORDER   = Border(
    left=Side(style="thin", color="AAAAAA"),
    right=Side(style="thin", color="AAAAAA"),
    top=Side(style="thin", color="AAAAAA"),
    bottom=Side(style="thin", color="AAAAAA"),
)

# Accent colour per language tab
ACCENT = {
    "Python":               "3572A5",  # Python blue
    "Go":                   "00ADD8",  # Go cyan
    "JS & TypeScript": "F7DF1E",  # JS yellow (dark text)
    "Java":                 "B07219",  # Java brown
    "PHP":                  "4F5D95",  # PHP indigo
    "C#":                   "178600",  # C# green
    "Ruby":                 "CC342D",  # Ruby red
    "Rust":                 "DEA584",  # Rust orange
    "Dart-Flutter":         "00B4AB",  # Dart teal
}

# Tab colours (need to be 6-char RGB hex for openpyxl)
TAB_COLOUR = {
    "Python":               "3572A5",
    "Go":                   "00ADD8",
    "JS & TypeScript":      "F0C030",
    "Java":                 "B07219",
    "PHP":                  "4F5D95",
    "C#":                   "178600",
    "Ruby":                 "CC342D",
    "Rust":                 "DEA584",
    "Dart-Flutter":         "00B4AB",
}

COLS = ["#", "Rule ID", "Library / Bundle", "What It Detects", "Code Pattern Matched"]
COL_WIDTHS = [5, 38, 28, 45, 65]

def add_sheet(wb: openpyxl.Workbook, lang: str, rules: list):
    ws = wb.create_sheet(title=lang)
    ws.sheet_properties.tabColor = TAB_COLOUR.get(lang, "888888")
    ws.freeze_panes = "A3"

    accent = ACCENT.get(lang, "1F3864")
    header_fill = PatternFill("solid", fgColor=accent)
    # For JS (yellow), use dark text
    hdr_font_color = "000000" if lang == "JS & TypeScript" else "FFFFFF"
    hdr_font = Font(bold=True, color=hdr_font_color, name="Calibri", size=11)

    # Title row
    ws.merge_cells("A1:E1")
    title_cell = ws["A1"]
    title_cell.value = f"cbom-scanner  —  {lang}  Detection Rules  ({len(rules)} rules)"
    title_cell.font = Font(bold=True, name="Calibri", size=13, color="FFFFFF")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    title_cell.fill = PatternFill("solid", fgColor="1F3864")
    ws.row_dimensions[1].height = 24

    # Header row
    for col_idx, (header, width) in enumerate(zip(COLS, COL_WIDTHS), start=1):
        cell = ws.cell(row=2, column=col_idx, value=header)
        cell.font = hdr_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[2].height = 18

    # Data rows
    for row_idx, (rule_id, bundle, description, pattern) in enumerate(rules, start=1):
        excel_row = row_idx + 2
        fill = ALT_FILL if row_idx % 2 == 0 else WHITE_FILL
        values = [row_idx, rule_id, bundle, description, pattern]
        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=excel_row, column=col_idx, value=val)
            cell.font = BODY_FONT
            cell.fill = fill
            cell.border = THIN_BORDER
            cell.alignment = Alignment(
                horizontal="center" if col_idx == 1 else "left",
                vertical="center",
                wrap_text=True,
            )
        ws.row_dimensions[excel_row].height = 30

    # Autofilter on header
    ws.auto_filter.ref = f"A2:E{len(rules)+2}"

    # No sheet protection
    ws.protection.sheet = False


def make_single_workbook(lang: str, rules: list) -> openpyxl.Workbook:
    """Create a standalone workbook for one language."""
    wb = openpyxl.Workbook()
    ws = wb.active

    accent = ACCENT.get(lang, "1F3864")
    hdr_font_color = "000000" if lang == "JS & TypeScript" else "FFFFFF"
    hdr_font = Font(bold=True, color=hdr_font_color, name="Calibri", size=11)
    header_fill = PatternFill("solid", fgColor=accent)

    ws.title = lang[:31]   # Excel sheet name limit
    ws.freeze_panes = "A3"

    # Title row
    ws.merge_cells("A1:E1")
    tc = ws["A1"]
    tc.value = f"cbom-scanner  —  {lang}  Detection Rules  ({len(rules)} rules)"
    tc.font = Font(bold=True, name="Calibri", size=14, color="FFFFFF")
    tc.alignment = Alignment(horizontal="center", vertical="center")
    tc.fill = PatternFill("solid", fgColor="1F3864")
    ws.row_dimensions[1].height = 28

    # Header row
    for col_idx, (header, width) in enumerate(zip(COLS, COL_WIDTHS), start=1):
        cell = ws.cell(row=2, column=col_idx, value=header)
        cell.font = hdr_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[2].height = 20

    # Data rows
    for row_idx, (rule_id, bundle, description, pattern) in enumerate(rules, start=1):
        excel_row = row_idx + 2
        fill = ALT_FILL if row_idx % 2 == 0 else WHITE_FILL
        for col_idx, val in enumerate([row_idx, rule_id, bundle, description, pattern], start=1):
            cell = ws.cell(row=excel_row, column=col_idx, value=val)
            cell.font = BODY_FONT
            cell.fill = fill
            cell.border = THIN_BORDER
            cell.alignment = Alignment(
                horizontal="center" if col_idx == 1 else "left",
                vertical="center",
                wrap_text=True,
            )
        ws.row_dimensions[excel_row].height = 32

    ws.auto_filter.ref = f"A2:E{len(rules)+2}"
    ws.protection.sheet = False
    return wb


def main():
    desktop = os.path.join(os.path.expanduser("~"), "Desktop")

    # ── One combined workbook, one sheet per language ──────────────────────
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove blank default sheet

    for lang, rules in RULES.items():
        accent = ACCENT.get(lang, "1F3864")
        hdr_font_color = "000000" if lang == "JS & TypeScript" else "FFFFFF"
        hdr_font = Font(bold=True, color=hdr_font_color, name="Calibri", size=11)
        header_fill = PatternFill("solid", fgColor=accent)

        ws = wb.create_sheet(title=lang[:31])
        ws.sheet_properties.tabColor = TAB_COLOUR.get(lang, "888888")
        ws.freeze_panes = "A3"

        # Title row
        ws.merge_cells("A1:E1")
        tc = ws["A1"]
        tc.value = f"cbom-scanner  —  {lang}  Detection Rules  ({len(rules)} rules)"
        tc.font = Font(bold=True, name="Calibri", size=14, color="FFFFFF")
        tc.alignment = Alignment(horizontal="center", vertical="center")
        tc.fill = PatternFill("solid", fgColor="1F3864")
        ws.row_dimensions[1].height = 28

        # Header row
        for col_idx, (header, width) in enumerate(zip(COLS, COL_WIDTHS), start=1):
            cell = ws.cell(row=2, column=col_idx, value=header)
            cell.font = hdr_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = THIN_BORDER
            ws.column_dimensions[get_column_letter(col_idx)].width = width
        ws.row_dimensions[2].height = 20

        # Data rows
        for row_idx, (rule_id, bundle, description, pattern) in enumerate(rules, start=1):
            excel_row = row_idx + 2
            fill = ALT_FILL if row_idx % 2 == 0 else WHITE_FILL
            for col_idx, val in enumerate([row_idx, rule_id, bundle, description, pattern], start=1):
                cell = ws.cell(row=excel_row, column=col_idx, value=val)
                cell.font = BODY_FONT
                cell.fill = fill
                cell.border = THIN_BORDER
                cell.alignment = Alignment(
                    horizontal="center" if col_idx == 1 else "left",
                    vertical="center",
                    wrap_text=True,
                )
            ws.row_dimensions[excel_row].height = 32

        ws.auto_filter.ref = f"A2:E{len(rules)+2}"
        ws.protection.sheet = False
        print(f"  [{len(rules):3d} rules]  {lang}")

    out = os.path.join(desktop, "cbom_scanner_rules_reference.xlsx")
    wb.save(out)
    print(f"\nSaved: {out}")


if __name__ == "__main__":
    main()
