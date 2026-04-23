"""
Generates a formatted CBOM security review Excel report from the sample CBOM data.
"""

import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter

OUTPUT_FILE = r"C:\Users\Nochu\Desktop\cbom_security_report.xlsx"

# ── Colour palette ────────────────────────────────────────────────────────────
CLR_HEADER_BG   = "1F3864"   # dark navy
CLR_HEADER_FG   = "FFFFFF"
CLR_PASS_BG     = "E2EFDA"   # light green
CLR_PASS_FG     = "375623"
CLR_REVIEW_BG   = "FFF2CC"   # amber
CLR_REVIEW_FG   = "7F6000"
CLR_ROW_ALT     = "F2F2F2"   # light grey alternate row
CLR_BORDER      = "B8CCE4"

# ── Thin border helper ────────────────────────────────────────────────────────
def thin_border():
    s = Side(style="thin", color=CLR_BORDER)
    return Border(left=s, right=s, top=s, bottom=s)

# ── Data ──────────────────────────────────────────────────────────────────────
ROWS = [
    # (asset, type, location_file, line, key_len, mode, status, severity, finding_id, notes)
    (
        "HS256",
        "JWT Signing Algorithm (Symmetric MAC)",
        "lib/custom_code/actions/homepage_nav_service.dart",
        1331, "-", "-",
        "REVIEW NEEDED", "Medium", "CBOM-JWT-002",
        (
            "Manual review required — static analysis cannot determine the signing "
            "secret length, how it was generated, or how widely it is shared.\n\n"
            "Please verify all three of the following:\n"
            "  1. Secret length: The signing secret must be at least 256 bits "
            "(32 bytes). A shorter secret can be cracked offline using brute-force tools.\n"
            "  2. Random generation: The secret must be generated using a "
            "cryptographically secure random generator (CSPRNG), not a human-chosen "
            "password or hardcoded string.\n"
            "  3. Key sharing: The same secret should not be shared across many "
            "services. Every service that holds the secret can both verify AND forge "
            "tokens — so limit sharing to the minimum necessary.\n\n"
            "HS256 is a symmetric signing algorithm — the same key signs and verifies. "
            "This is acceptable when used correctly. You can continue using a symmetric "
            "signing algorithm. Consider upgrading to HS512 for a stronger HMAC "
            "(example in Dart: JWTAlgorithm.HS512).\n\n"
            "If the secret is shared across more than 2–3 services, consider switching "
            "to an asymmetric algorithm such as RS256 (RSA) or ES256 (ECDSA), where "
            "only one server holds the private signing key and all others only hold "
            "the public verification key."
        ),
    ),
    (
        "AES-256",
        "Symmetric Block Cipher",
        "lib/utils/payload_utils.dart",
        13, "256 bits", "-",
        "PASS", "-", "-",
        "AES-256 is a strong, NIST-approved symmetric cipher. No issues found.",
    ),
    (
        "AES-256",
        "Symmetric Block Cipher",
        "lib/utils/payload_utils.dart",
        26, "256 bits", "-",
        "PASS", "-", "-",
        "AES-256 is a strong, NIST-approved symmetric cipher. No issues found.",
    ),
    (
        "AES-256-CBC",
        "Symmetric Block Cipher",
        "lib/backend/api_requests/encryption.dart",
        134, "256 bits", "CBC",
        "PASS", "-", "-",
        "AES-256-CBC is secure. Ensure a unique random IV is used for every encryption "
        "operation and is stored alongside the ciphertext.",
    ),
    (
        "HS256",
        "JWT Signing Algorithm (Symmetric MAC)",
        "lib/custom_code/actions/get_jwt_token.dart",
        31, "-", "-",
        "REVIEW NEEDED", "Medium", "CBOM-JWT-002",
        (
            "Manual review required — static analysis cannot determine the signing "
            "secret length, how it was generated, or how widely it is shared.\n\n"
            "Please verify all three of the following:\n"
            "  1. Secret length: The signing secret must be at least 256 bits "
            "(32 bytes). A shorter secret can be cracked offline using brute-force tools.\n"
            "  2. Random generation: The secret must be generated using a "
            "cryptographically secure random generator (CSPRNG), not a human-chosen "
            "password or hardcoded string.\n"
            "  3. Key sharing: The same secret should not be shared across many "
            "services. Every service that holds the secret can both verify AND forge "
            "tokens — so limit sharing to the minimum necessary.\n\n"
            "HS256 is a symmetric signing algorithm — the same key signs and verifies. "
            "This is acceptable when used correctly. You can continue using a symmetric "
            "signing algorithm. Consider upgrading to HS512 for a stronger HMAC "
            "(example in Dart: JWTAlgorithm.HS512).\n\n"
            "If the secret is shared across more than 2–3 services, consider switching "
            "to an asymmetric algorithm such as RS256 (RSA) or ES256 (ECDSA), where "
            "only one server holds the private signing key and all others only hold "
            "the public verification key."
        ),
    ),
    (
        "AES-256-CBC",
        "Symmetric Block Cipher",
        "lib/backend/api_requests/encryption.dart",
        120, "256 bits", "CBC",
        "PASS", "-", "-",
        "AES-256-CBC is secure. Ensure a unique random IV is used for every encryption "
        "operation and is stored alongside the ciphertext.",
    ),
    (
        "Random.secure",
        "Cryptographically Secure PRNG (CSPRNG)",
        "lib/backend/firebase/firebase_config.dart",
        60, "-", "-",
        "PASS", "-", "-",
        "Dart's Random.secure() is a CSPRNG backed by the OS entropy source. "
        "This is the correct function to use for cryptographic key/nonce generation.",
    ),
    (
        "AES-256-CBC",
        "Symmetric Block Cipher",
        "personal_loan/lib/state/personal_loan_view_model.dart",
        2884, "256 bits", "CBC",
        "PASS", "-", "-",
        "AES-256-CBC is secure. Ensure a unique random IV is used for every encryption "
        "operation and is stored alongside the ciphertext.",
    ),
    (
        "AES-256-CBC",
        "Symmetric Block Cipher",
        "pl_multi_lender/lib/feature/pl_multi_lender_common_function/...",
        722, "256 bits", "CBC",
        "PASS", "-", "-",
        "AES-256-CBC is secure. Ensure a unique random IV is used for every encryption "
        "operation and is stored alongside the ciphertext.",
    ),
    (
        "AES-256-CBC",
        "Symmetric Block Cipher",
        "homeloannewmain/lib/data/localdb/localdbHelper.dart",
        43, "256 bits", "CBC",
        "PASS", "-", "-",
        "AES-256-CBC is secure. Ensure a unique random IV is used for every encryption "
        "operation and is stored alongside the ciphertext.",
    ),
    (
        "AES-256-CBC",
        "Symmetric Block Cipher",
        "lib/backend/api_requests/encryption.dart",
        76, "256 bits", "CBC",
        "PASS", "-", "-",
        "AES-256-CBC is secure. Ensure a unique random IV is used for every encryption "
        "operation and is stored alongside the ciphertext.",
    ),
]

HEADERS = [
    "#",
    "Asset Name",
    "Asset Type",
    "File Location",
    "Line",
    "Key Length",
    "Mode",
    "Status",
    "Severity",
    "Finding ID",
    "Notes & Review Guidance",
]

COL_WIDTHS = [4, 16, 30, 60, 7, 12, 8, 16, 10, 16, 80]

# ── Build workbook ────────────────────────────────────────────────────────────
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "CBOM Security Report"

# Title row
ws.merge_cells("A1:K1")
title_cell = ws["A1"]
title_cell.value = "CBOM Security Report — abcd-frontend"
title_cell.font = Font(name="Calibri", size=14, bold=True, color=CLR_HEADER_FG)
title_cell.fill = PatternFill("solid", fgColor=CLR_HEADER_BG)
title_cell.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 28

# Subtitle row
ws.merge_cells("A2:K2")
sub_cell = ws["A2"]
sub_cell.value = (
    "Generated by cbom-scanner v0.1.0  |  Scanned: 2026-04-22  |  "
    "Format: CycloneDX 1.6  |  Language: Dart / Flutter"
)
sub_cell.font = Font(name="Calibri", size=10, italic=True, color="595959")
sub_cell.fill = PatternFill("solid", fgColor="D9E2F3")
sub_cell.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[2].height = 18

# Blank spacer
ws.row_dimensions[3].height = 6

# Header row
header_row = 4
for col_idx, (header, width) in enumerate(zip(HEADERS, COL_WIDTHS), start=1):
    cell = ws.cell(row=header_row, column=col_idx, value=header)
    cell.font = Font(name="Calibri", size=10, bold=True, color=CLR_HEADER_FG)
    cell.fill = PatternFill("solid", fgColor=CLR_HEADER_BG)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = thin_border()
    ws.column_dimensions[get_column_letter(col_idx)].width = width

ws.row_dimensions[header_row].height = 22

# Data rows
for row_num, row_data in enumerate(ROWS, start=1):
    (asset, atype, location, line, key_len, mode,
     status, severity, finding_id, notes) = row_data

    excel_row = header_row + row_num
    is_review = (status == "REVIEW NEEDED")

    row_bg = CLR_REVIEW_BG if is_review else (CLR_PASS_BG if status == "PASS" else CLR_ROW_ALT)

    values = [row_num, asset, atype, location, line,
              key_len, mode, status, severity, finding_id, notes]

    for col_idx, value in enumerate(values, start=1):
        cell = ws.cell(row=excel_row, column=col_idx, value=value)
        cell.font = Font(name="Calibri", size=9)
        cell.border = thin_border()
        cell.fill = PatternFill("solid", fgColor=row_bg)

        # Notes column: wrap text, left-aligned
        if col_idx == 11:
            cell.alignment = Alignment(
                vertical="top", wrap_text=True, horizontal="left"
            )
        # Status column: colour + bold
        elif col_idx == 8:
            if is_review:
                cell.font = Font(name="Calibri", size=9, bold=True, color=CLR_REVIEW_FG)
                cell.value = "⚠  REVIEW NEEDED"
            else:
                cell.font = Font(name="Calibri", size=9, bold=True, color=CLR_PASS_FG)
                cell.value = "✅  PASS"
            cell.alignment = Alignment(horizontal="center", vertical="center")
        # Severity column
        elif col_idx == 9:
            if severity == "Medium":
                cell.font = Font(name="Calibri", size=9, bold=True, color="7F6000")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        # Line number
        elif col_idx == 5:
            cell.alignment = Alignment(horizontal="center", vertical="center")
        # Key length / mode
        elif col_idx in (6, 7):
            cell.alignment = Alignment(horizontal="center", vertical="center")
        # Asset name: bold
        elif col_idx == 2:
            cell.font = Font(name="Calibri", size=9, bold=True)
            cell.alignment = Alignment(vertical="center")
        else:
            cell.alignment = Alignment(vertical="center", wrap_text=False)

    # Auto-height: estimate by line count in notes
    note_lines = notes.count("\n") + max(1, len(notes) // 95)
    ws.row_dimensions[excel_row].height = max(20, note_lines * 14 + 6)

# Freeze panes below header
ws.freeze_panes = "A5"

# Auto-filter on header row
ws.auto_filter.ref = f"A{header_row}:K{header_row}"

# Explicitly ensure no sheet protection — all cells fully editable
ws.protection.sheet = False
ws.protection.enable()   # calling enable() then immediately disabling ensures clean state
ws.protection.sheet = False

# ── Legend sheet ──────────────────────────────────────────────────────────────
ls = wb.create_sheet("Legend")
ls.column_dimensions["A"].width = 22
ls.column_dimensions["B"].width = 70

legend_title = ls.cell(row=1, column=1, value="Status Legend")
legend_title.font = Font(name="Calibri", size=12, bold=True, color=CLR_HEADER_FG)
legend_title.fill = PatternFill("solid", fgColor=CLR_HEADER_BG)
ls.merge_cells("A1:B1")
ls["B1"].fill = PatternFill("solid", fgColor=CLR_HEADER_BG)

legend_data = [
    ("✅  PASS",         "Asset uses a strong, approved algorithm with no known weaknesses detected."),
    ("⚠  REVIEW NEEDED","Static analysis flagged a pattern that requires manual verification. "
                         "The asset may be fine if configured correctly — see the Notes column."),
    ("Medium Severity",  "The finding is a moderate concern. Not an immediate critical risk, "
                         "but must be reviewed and resolved before production deployment."),
]

for r, (label, desc) in enumerate(legend_data, start=2):
    a = ls.cell(row=r, column=1, value=label)
    b = ls.cell(row=r, column=2, value=desc)
    bg = CLR_REVIEW_BG if "REVIEW" in label or "Medium" in label else CLR_PASS_BG
    a.fill = PatternFill("solid", fgColor=bg)
    b.fill = PatternFill("solid", fgColor=bg)
    a.font = Font(name="Calibri", size=9, bold=True)
    b.font = Font(name="Calibri", size=9)
    a.border = thin_border()
    b.border = thin_border()
    b.alignment = Alignment(wrap_text=True, vertical="top")
    ls.row_dimensions[r].height = 32

wb.save(OUTPUT_FILE)
print(f"Report saved -> {OUTPUT_FILE}")
